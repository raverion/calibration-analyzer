# excel_charts.py
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, Alignment, PatternFill

from utils import CHANNEL_COLORS  # Import shared constant

def apply_channel_colors_to_results(excel_file, df_results, unit, color_assignments):
    """
    Apply channel colors to the Test Results sheet based on the color assignments
    from the tolerance charts.
    """
    from openpyxl.styles import Font, PatternFill
    
    wb = load_workbook(excel_file)
    ws = wb['Test Results']
    
    # Columns A to M (1 to 13)
    color_columns = list(range(1, 14))
    
    # Iterate through data rows (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        if row_idx >= len(df_results):
            break
        
        # Get the data for this row from the dataframe
        row_data = df_results.iloc[row_idx]
        channel = row_data['Channel']
        io_type = row_data['I/O Type']
        test_value = row_data[f'Test Value [{unit}]']
        range_setting = row_data['Range Setting']
        
        # Look up the color for this combination (range_setting is used as-is, including 'N/A')
        key = (channel, io_type, test_value, range_setting)
        color = color_assignments.get(key)
        
        if color:
            # Apply color to columns A through M
            for cell in row:
                if cell.column in color_columns:
                    # Preserve existing formatting but change font color
                    current_font = cell.font
                    cell.font = Font(
                        name=current_font.name,
                        size=current_font.size,
                        bold=current_font.bold,
                        italic=current_font.italic,
                        color=color
                    )
    
    wb.save(excel_file)
    print("✓ Channel colors applied to Test Results sheet")

def create_tolerance_charts(excel_file, df_results, unit):
    """
    Create tolerance charts showing limits, reference value, mean, and mean±2σ for each 
    test value + range setting combination.
    
    Returns: Dictionary mapping (channel, io_type, test_value, range_setting) to color
    """
    print("\nCreating Tolerance charts...")
    
    wb = load_workbook(excel_file)
    
    if 'Tolerance Charts' in wb.sheetnames:
        del wb['Tolerance Charts']
    chart_sheet = wb.create_sheet('Tolerance Charts')
    
    # Get unique combinations
    unique_combinations = df_results.groupby([f'Test Value [{unit}]', 'Range Setting', 'I/O Type']).size().reset_index()
    unique_combinations = unique_combinations.sort_values([f'Test Value [{unit}]', 'Range Setting', 'I/O Type'])
    
    charts_per_row = 2
    chart_width = 26
    chart_height = 36
    
    # Pleasant, muted color palette for channels (same color for mean, +2σ, -2σ)
    channel_colors = [
        '4472C4',  # Muted blue
        'C45B5B',  # Muted red
        '70AD47',  # Muted green
        'ED7D31',  # Muted orange
        '7B7B7B',  # Gray
        '9E5ECE',  # Muted purple
        '43A6A2',  # Muted teal
        'C4A24E',  # Muted gold
        '5B9BC4',  # Steel blue
        'A85B5B',  # Dusty rose
        '5BAF7B',  # Sea green
        'C47B4E',  # Terracotta
        '6B6BAF',  # Muted indigo
        '8B6BAF',  # Muted violet
        '4EAFAF',  # Turquoise
        'AF8B4E',  # Bronze
    ]
    
    # Track color assignments for each channel within each chart context
    # Key: (channel, io_type, test_value, range_setting) -> color
    color_assignments = {}
    
    chart_idx = 0
    
    from openpyxl.styles import Font, Alignment, PatternFill
    
    for _, combo_row in unique_combinations.iterrows():
        test_value = combo_row[f'Test Value [{unit}]']
        range_setting = combo_row['Range Setting']
        io_type = combo_row['I/O Type']
        
        mask = (
            (df_results[f'Test Value [{unit}]'] == test_value) &
            (df_results['Range Setting'] == range_setting) &
            (df_results['I/O Type'] == io_type)
        )
        test_data = df_results[mask].copy()
        test_data = test_data.sort_values('Channel')
        
        if len(test_data) == 0:
            continue
        
        col_offset = (chart_idx % charts_per_row) * chart_width
        row_offset = (chart_idx // charts_per_row) * chart_height
        
        range_display = f", Range: {range_setting}" if range_setting != 'N/A' else ""
        chart_title = f"Test: {test_value} {unit}{range_display} ({io_type})"
        
        title_row = row_offset + 1
        chart_sheet.cell(title_row, col_offset + 1).value = chart_title
        title_cell = chart_sheet.cell(title_row, col_offset + 1)
        title_cell.font = Font(bold=True, size=11, color='1F4E78')
        title_cell.alignment = Alignment(horizontal='left')
        
        channels = test_data['Channel'].tolist()
        num_channels = len(channels)
        lower_limits = test_data[f'Lower Limit [{unit}]'].tolist()
        upper_limits = test_data[f'Upper Limit [{unit}]'].tolist()
        reference_values = test_data[f'Reference Value [{unit}]'].tolist()
        means = test_data[f'Mean [{unit}]'].tolist()
        lower_2sigma = (test_data[f'Mean [{unit}]'] - 2*test_data[f'StdDev [{unit}]']).tolist()
        upper_2sigma = (test_data[f'Mean [{unit}]'] + 2*test_data[f'StdDev [{unit}]']).tolist()
        mean_checks = test_data['Mean Check'].tolist()
        mean_2sigma_checks = test_data['Mean±2σ Check'].tolist()
        
        # Get limit values (same for all channels in this chart)
        ref_val = reference_values[0]
        ll_val = lower_limits[0]
        ul_val = upper_limits[0]
        
        data_start_row = row_offset + 3
        data_start_col = col_offset + 1
        
        # Write headers
        headers = ['Channel', 'Lower Limit', 'Reference', 'Upper Limit', 'Mean', 'Mean-2σ', 'Mean+2σ', 'Mean Check', 'Mean±2σ Check']
        for h_idx, header in enumerate(headers):
            chart_sheet.cell(data_start_row, data_start_col + h_idx).value = header
            chart_sheet.cell(data_start_row, data_start_col + h_idx).font = Font(bold=True, size=9)
        
        # Write data with color-coded fonts
        for i, channel in enumerate(channels):
            row = data_start_row + i + 1
            color = channel_colors[i % len(channel_colors)]
            
            # Store color assignment
            color_assignments[(channel, io_type, test_value, range_setting)] = color
            
            # Apply color to all cells in this row
            cell_font = Font(size=9, color=color)
            
            chart_sheet.cell(row, data_start_col, channel).font = cell_font
            chart_sheet.cell(row, data_start_col + 1, lower_limits[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 1).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 2, reference_values[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 2).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 3, upper_limits[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 3).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 4, means[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 4).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 5, lower_2sigma[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 5).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 6, upper_2sigma[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 6).number_format = '0.000000'
            
            # Mean Check column with PASS/FAIL formatting
            mean_check_cell = chart_sheet.cell(row, data_start_col + 7, mean_checks[i])
            if mean_checks[i] == 'PASS':
                mean_check_cell.font = Font(size=9, color='006100', bold=True)
                mean_check_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                mean_check_cell.font = Font(size=9, color='9C0006', bold=True)
                mean_check_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Mean±2σ Check column with PASS/FAIL formatting
            mean_2sigma_check_cell = chart_sheet.cell(row, data_start_col + 8, mean_2sigma_checks[i])
            if mean_2sigma_checks[i] == 'PASS':
                mean_2sigma_check_cell.font = Font(size=9, color='006100', bold=True)
                mean_2sigma_check_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                mean_2sigma_check_cell.font = Font(size=9, color='9C0006', bold=True)
                mean_2sigma_check_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Create scatter chart
        chart = ScatterChart()
        chart.title = chart_title
        chart.style = 10
        
        # Set axis titles (non-bold is default for axis titles)
        chart.x_axis.title = "Channel"
        chart.y_axis.title = f"Measurement [{unit}]"
        
        # Chart size - make it larger for better visibility
        chart.height = 14
        chart.width = 18
        
        # Remove legend
        chart.legend = None
        
        # Remove gridlines
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
        
        # Calculate Y-axis limits with padding
        all_values = lower_limits + upper_limits + reference_values + means + lower_2sigma + upper_2sigma
        y_min = min(all_values)
        y_max = max(all_values)
        y_range = y_max - y_min if y_max != y_min else abs(y_max) * 0.1 or 0.1
        y_padding = y_range * 0.20  # 20% padding
        
        chart.y_axis.scaling.min = y_min - y_padding
        chart.y_axis.scaling.max = y_max + y_padding
        
        # Set X-axis to properly scale for the number of channels
        x_min_val = min(channels)
        x_max_val = max(channels)
        x_padding_left = max(0.8, (x_max_val - x_min_val) * 0.1)
        x_padding_right = max(0.8, (x_max_val - x_min_val) * 0.1)
        chart.x_axis.scaling.min = x_min_val - x_padding_left
        chart.x_axis.scaling.max = x_max_val + x_padding_right
        chart.x_axis.majorUnit = 1  # Show each channel number
        
        # Reference for x values (channel numbers)
        xvalues = Reference(chart_sheet, min_col=data_start_col, min_row=data_start_row+1, max_row=data_start_row+num_channels)
        
        # Lower limit line (dark red, dashed)
        lower_series = Series(Reference(chart_sheet, min_col=data_start_col+1, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                            xvalues, title="Lower Limit")
        lower_series.marker = Marker('none')
        lower_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="8B0000", w=12700, prstDash='dash'))
        chart.series.append(lower_series)
        
        # Reference value line (dark green, solid)
        ref_series = Series(Reference(chart_sheet, min_col=data_start_col+2, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                           xvalues, title="Reference")
        ref_series.marker = Marker('none')
        ref_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="2E7D32", w=12700))
        chart.series.append(ref_series)
        
        # Upper limit line (dark red, dashed)
        upper_series = Series(Reference(chart_sheet, min_col=data_start_col+3, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                            xvalues, title="Upper Limit")
        upper_series.marker = Marker('none')
        upper_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="8B0000", w=12700, prstDash='dash'))
        chart.series.append(upper_series)
        
        # Add series for each channel with consistent colors
        for i, channel in enumerate(channels):
            color = channel_colors[i % len(channel_colors)]
            channel_row = data_start_row + 1 + i
            
            # Mean (diamond marker - smaller size)
            mean_series = Series(Reference(chart_sheet, min_col=data_start_col+4, min_row=channel_row, max_row=channel_row), 
                               Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                               title=f"CH{channel} Mean")
            mean_series.marker = Marker('diamond', size=6)
            mean_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color))
            mean_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(mean_series)
            
            # Mean-2σ (horizontal line marker - thinner)
            lower_2s_series = Series(Reference(chart_sheet, min_col=data_start_col+5, min_row=channel_row, max_row=channel_row), 
                                   Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                                   title=f"CH{channel} -2σ")
            lower_2s_series.marker = Marker('dash', size=8)
            lower_2s_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color, w=12700))
            lower_2s_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(lower_2s_series)
            
            # Mean+2σ (horizontal line marker - thinner)
            upper_2s_series = Series(Reference(chart_sheet, min_col=data_start_col+6, min_row=channel_row, max_row=channel_row), 
                                   Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                                   title=f"CH{channel} +2σ")
            upper_2s_series.marker = Marker('dash', size=8)
            upper_2s_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color, w=12700))
            upper_2s_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(upper_2s_series)
        
        # Position chart (moved further right to accommodate check columns)
        chart_cell = chart_sheet.cell(row_offset + 4, col_offset + 12)
        chart.anchor = chart_cell.coordinate
        chart_sheet.add_chart(chart)
        
        print(f"  Created chart for {chart_title}")
        chart_idx += 1
    
    wb.save(excel_file)
    print("✓ Tolerance charts added to workbook")
    
    return color_assignments
