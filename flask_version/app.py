"""
Flask Web Application for Measurement Data Analysis
Converts the tkinter-based GUI to a web interface while preserving all backend functionality.
"""

import os
import json
import uuid
import shutil
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for

import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font

# Import from local modules (same as original)
from parsers import (
    parse_filename,
    get_unit_from_files,
    scan_text_file_for_measurement_types,
    parse_text_file,
    extract_equipment_name
)
from excel_charts import create_tolerance_charts, apply_channel_colors_to_results
from html_report import create_html_report
from utils import get_versioned_filename

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Configuration
UPLOAD_FOLDER = Path(__file__).parent / 'uploads'
OUTPUT_FOLDER = Path(__file__).parent / 'outputs'
UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
app.config['OUTPUT_FOLDER'] = str(OUTPUT_FOLDER)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max upload


def get_session_folder():
    """Get or create a unique folder for this session's uploads."""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    
    session_folder = UPLOAD_FOLDER / session['session_id']
    session_folder.mkdir(exist_ok=True)
    return session_folder


def get_output_folder():
    """Get or create a unique folder for this session's outputs."""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    
    output_folder = OUTPUT_FOLDER / session['session_id']
    output_folder.mkdir(exist_ok=True)
    return output_folder


@app.route('/')
def index():
    """Landing page with mode selection."""
    return render_template('index.html')


@app.route('/equipment-report')
def equipment_report():
    """Equipment-specific report page - upload files."""
    # Clear previous session data
    session.pop('files_info', None)
    session.pop('measurement_types', None)
    session.pop('test_configs', None)
    return render_template('upload.html')


@app.route('/comparison-report')
def comparison_report():
    """Cross-equipment comparison report - upload Excel reports."""
    # Clear previous session data for comparison
    session.pop('comparison_files', None)
    session.pop('comparison_data', None)
    return render_template('comparison_upload.html')


@app.route('/api/upload-comparison', methods=['POST'])
def upload_comparison_files():
    """Handle Excel report uploads for cross-equipment comparison."""
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files selected'}), 400
    
    session_folder = get_session_folder()
    comparison_folder = session_folder / 'comparison'
    comparison_folder.mkdir(exist_ok=True)
    
    # Clear previous comparison uploads
    for old_file in comparison_folder.iterdir():
        old_file.unlink()
    
    uploaded_files = []
    validation_errors = []
    equipment_types = set()
    test_configs_by_file = {}
    
    for file in files:
        if file.filename and file.filename.lower().endswith('.xlsx'):
            filename = Path(file.filename).name
            file_path = comparison_folder / filename
            file.save(str(file_path))
            
            # Validate the Excel file
            validation_result = validate_equipment_report(file_path)
            
            if validation_result['valid']:
                uploaded_files.append({
                    'filename': filename,
                    'equipment_type': validation_result['equipment_type'],
                    'sample_id': validation_result['sample_id'],
                    'unit': validation_result['unit'],
                    'test_values': validation_result['test_values'],
                    'channels': validation_result['channels'],
                    'io_types': validation_result['io_types']
                })
                equipment_types.add(validation_result['equipment_type'])
                test_configs_by_file[filename] = validation_result['test_values']
            else:
                validation_errors.append({
                    'filename': filename,
                    'error': validation_result['error']
                })
                # Remove invalid file
                file_path.unlink()
    
    if not uploaded_files:
        return jsonify({
            'error': 'No valid Excel reports found',
            'validation_errors': validation_errors
        }), 400
    
    # Check for multiple equipment types
    warnings = []
    if len(equipment_types) > 1:
        warnings.append({
            'type': 'multiple_equipment_types',
            'message': f"Multiple equipment types detected: {', '.join(sorted(equipment_types))}. Please upload reports from the same equipment type only."
        })
    
    # Check for mismatched test values
    if len(uploaded_files) > 1:
        first_file = uploaded_files[0]
        first_test_values = set(first_file['test_values'])
        
        for file_info in uploaded_files[1:]:
            file_test_values = set(file_info['test_values'])
            if file_test_values != first_test_values:
                missing = first_test_values - file_test_values
                extra = file_test_values - first_test_values
                msg_parts = []
                if missing:
                    msg_parts.append(f"missing test values: {missing}")
                if extra:
                    msg_parts.append(f"extra test values: {extra}")
                warnings.append({
                    'type': 'mismatched_test_values',
                    'message': f"File '{file_info['filename']}' has different test values ({', '.join(msg_parts)})"
                })
    
    # Store in session
    session['comparison_files'] = uploaded_files
    session['comparison_warnings'] = warnings
    session['comparison_validation_errors'] = validation_errors
    
    return jsonify({
        'success': True,
        'files_count': len(uploaded_files),
        'files': uploaded_files,
        'equipment_types': list(equipment_types),
        'warnings': warnings,
        'validation_errors': validation_errors
    })


def validate_equipment_report(file_path):
    """
    Validate that an Excel file is a valid equipment report generated by this application.
    
    Returns dict with:
    - valid: bool
    - error: str (if not valid)
    - equipment_type: str (e.g., '50910')
    - sample_id: str (e.g., '50910-001')
    - unit: str
    - test_values: list of test values
    - channels: list of channel numbers
    - io_types: list of I/O types
    """
    try:
        # Try to load the Excel file
        xl = pd.ExcelFile(file_path)
        
        # Check for required sheet
        if 'Test Results' not in xl.sheet_names:
            return {'valid': False, 'error': 'Missing "Test Results" sheet - not a valid equipment report'}
        
        # Load the Test Results sheet
        df = pd.read_excel(file_path, sheet_name='Test Results')
        
        # Check for required columns
        required_columns = ['Channel', 'I/O Type', 'Range Setting']
        
        # Find the unit from column names
        unit = None
        for col in df.columns:
            if 'Test Value [' in col:
                unit = col.split('[')[1].split(']')[0]
                break
        
        if unit is None:
            return {'valid': False, 'error': 'Could not determine measurement unit from columns'}
        
        expected_columns = [
            'Channel', 'I/O Type', 'Range Setting',
            f'Test Value [{unit}]', f'Reference Value [{unit}]', f'Tolerance [{unit}]',
            f'Lower Limit [{unit}]', f'Upper Limit [{unit}]',
            f'Mean [{unit}]', f'StdDev [{unit}]', f'Min [{unit}]', f'Max [{unit}]',
            'Samples', 'Mean Check', 'Mean±2σ Check'
        ]
        
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            return {'valid': False, 'error': f'Missing required columns: {", ".join(missing_columns)}'}
        
        # Extract equipment type and sample ID from filename
        filename = Path(file_path).stem
        # Try to parse equipment type (e.g., "50910" from "50910-001")
        parts = filename.split('-')
        if len(parts) >= 2:
            equipment_type = parts[0]
            sample_id = filename
        else:
            equipment_type = filename
            sample_id = filename
        
        # Extract test values, channels, and I/O types
        test_values = df[f'Test Value [{unit}]'].unique().tolist()
        channels = sorted(df['Channel'].unique().tolist())
        io_types = df['I/O Type'].unique().tolist()
        
        return {
            'valid': True,
            'equipment_type': equipment_type,
            'sample_id': sample_id,
            'unit': unit,
            'test_values': test_values,
            'channels': channels,
            'io_types': io_types
        }
        
    except Exception as e:
        return {'valid': False, 'error': f'Error reading file: {str(e)}'}


@app.route('/comparison-configure')
def comparison_configure():
    """Configuration page for cross-equipment comparison."""
    if 'comparison_files' not in session:
        return redirect(url_for('comparison_report'))
    
    return render_template('comparison_configure.html',
                          files=session.get('comparison_files', []),
                          warnings=session.get('comparison_warnings', []),
                          validation_errors=session.get('comparison_validation_errors', []))


@app.route('/api/process-comparison', methods=['POST'])
def process_comparison():
    """Process uploaded Excel reports for cross-equipment comparison."""
    if 'comparison_files' not in session:
        return jsonify({'error': 'No files uploaded'}), 400
    
    data = request.json
    selected_channels = data.get('channels', 'all')  # 'all', 'mean', or list of channel numbers
    selected_io_type = data.get('io_type', 'all')  # 'all', 'Input', or 'Output'
    group_by = data.get('group_by', 'sample')  # 'sample' or 'channel'
    equipment_model = data.get('equipment_model', '')  # e.g., "VIO2004"
    
    session_folder = get_session_folder()
    comparison_folder = session_folder / 'comparison'
    output_folder = get_output_folder()
    
    try:
        # Load all Excel files and combine data
        all_data = []
        unit = None
        equipment_type = None
        
        for file_info in session['comparison_files']:
            file_path = comparison_folder / file_info['filename']
            df = pd.read_excel(file_path, sheet_name='Test Results')
            
            # Get unit and equipment type from first file
            if unit is None:
                for col in df.columns:
                    if 'Test Value [' in col:
                        unit = col.split('[')[1].split(']')[0]
                        break
            
            if equipment_type is None:
                equipment_type = file_info['equipment_type']
            
            # Add sample ID column
            df['Sample ID'] = file_info['sample_id']
            df['Equipment Type'] = file_info['equipment_type']
            
            all_data.append(df)
        
        # Combine all data
        combined_df = pd.concat(all_data, ignore_index=True)
        
        # Handle NaN values in Range Setting (important for groupby operations)
        combined_df['Range Setting'] = combined_df['Range Setting'].fillna('N/A')
        
        # Filter by I/O type if specified
        if selected_io_type != 'all':
            combined_df = combined_df[combined_df['I/O Type'] == selected_io_type]
        
        # Calculate normalized error (Mean - Reference)
        combined_df['Error'] = combined_df[f'Mean [{unit}]'] - combined_df[f'Reference Value [{unit}]']
        combined_df['Error-2σ'] = (combined_df[f'Mean [{unit}]'] - 2*combined_df[f'StdDev [{unit}]']) - combined_df[f'Reference Value [{unit}]']
        combined_df['Error+2σ'] = (combined_df[f'Mean [{unit}]'] + 2*combined_df[f'StdDev [{unit}]']) - combined_df[f'Reference Value [{unit}]']
        
        # Build full equipment identifier
        # e.g., "VIO2004_50920" if model provided, else just "50920"
        if equipment_model:
            full_equipment_name = f"{equipment_model}_{equipment_type}"
        else:
            full_equipment_name = equipment_type
        
        # Generate comparison report
        html_file = create_comparison_html_report(
            combined_df, 
            unit, 
            output_folder,
            selected_channels,
            session['comparison_files'],
            group_by,
            full_equipment_name
        )
        
        session['comparison_output'] = {
            'html': os.path.basename(html_file),
            'equipment_type': equipment_type,
            'equipment_model': equipment_model
        }
        
        return jsonify({
            'success': True,
            'html_file': os.path.basename(html_file)
        })
        
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


def create_comparison_html_report(df, unit, output_folder, selected_channels, files_info, group_by='sample', equipment_type=None):
    """
    Create an interactive HTML report for cross-equipment comparison.
    
    Parameters:
    - df: Combined DataFrame with all equipment data
    - unit: Measurement unit
    - output_folder: Output directory
    - selected_channels: 'all', 'mean', or list of channel numbers
    - files_info: List of file info dicts
    - group_by: 'sample' (group by equipment sample) or 'channel' (group by channel number)
    - equipment_type: Equipment type/model for report title
    """
    from utils import CHANNEL_COLORS_HEX
    
    # Use equipment type for filename if available
    report_name = equipment_type if equipment_type else 'comparison'
    
    # Generate HTML filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_file = os.path.join(str(output_folder), f'{report_name}_comparison_{timestamp}.html')
    
    # Get unique test value + range + I/O type combinations
    unique_combinations = df.groupby([f'Test Value [{unit}]', 'Range Setting', 'I/O Type']).size().reset_index()
    unique_combinations = unique_combinations.sort_values([f'Test Value [{unit}]', 'Range Setting', 'I/O Type'])
    
    # Get unique samples and channels
    sample_ids = df['Sample ID'].unique().tolist()
    channel_ids = sorted(df['Channel'].unique().tolist())
    
    # Assign colors based on grouping mode
    if group_by == 'sample':
        # Colors represent equipment samples
        group_colors = {sample: CHANNEL_COLORS_HEX[i % len(CHANNEL_COLORS_HEX)] for i, sample in enumerate(sample_ids)}
        legend_items = sample_ids
    else:
        # Colors represent channels
        group_colors = {ch: CHANNEL_COLORS_HEX[i % len(CHANNEL_COLORS_HEX)] for i, ch in enumerate(channel_ids)}
        legend_items = [f"CH{ch}" for ch in channel_ids]
    
    # Build chart data for each combination
    charts_data = []
    
    for _, combo_row in unique_combinations.iterrows():
        test_value = combo_row[f'Test Value [{unit}]']
        range_setting = combo_row['Range Setting']
        io_type = combo_row['I/O Type']
        
        mask = (
            (df[f'Test Value [{unit}]'] == test_value) &
            (df['Range Setting'] == range_setting) &
            (df['I/O Type'] == io_type)
        )
        chart_data = df[mask].copy()
        
        if len(chart_data) == 0:
            continue
        
        # Get tolerance (should be same for all entries in this chart)
        tolerance = chart_data[f'Tolerance [{unit}]'].iloc[0]
        
        range_display = f", Range: {range_setting}" if range_setting != 'N/A' else ""
        chart_title = f"Test: {test_value} {unit}{range_display} ({io_type})"
        
        # Prepare data for chart based on grouping mode
        chart_info = {
            'title': chart_title,
            'test_value': test_value,
            'range_setting': range_setting,
            'io_type': io_type,
            'tolerance': tolerance,
            'group_by': group_by,
            'groups': []
        }
        
        if group_by == 'sample':
            # Group by equipment sample (original behavior)
            for sample_id in sample_ids:
                sample_data = chart_data[chart_data['Sample ID'] == sample_id]
                if len(sample_data) == 0:
                    continue
                
                items_data = []
                for _, row in sample_data.iterrows():
                    items_data.append({
                        'label': f"CH{int(row['Channel'])}",
                        'channel': int(row['Channel']),
                        'sample_id': sample_id,
                        'error': row['Error'],
                        'error_minus_2sigma': row['Error-2σ'],
                        'error_plus_2sigma': row['Error+2σ'],
                        'mean_check': row['Mean Check'],
                        'sigma_check': row['Mean±2σ Check']
                    })
                
                chart_info['groups'].append({
                    'group_id': sample_id,
                    'group_label': sample_id,
                    'color': group_colors[sample_id],
                    'items': items_data
                })
        else:
            # Group by channel number
            for channel_id in channel_ids:
                channel_data = chart_data[chart_data['Channel'] == channel_id]
                if len(channel_data) == 0:
                    continue
                
                items_data = []
                for _, row in channel_data.iterrows():
                    items_data.append({
                        'label': row['Sample ID'],
                        'channel': int(row['Channel']),
                        'sample_id': row['Sample ID'],
                        'error': row['Error'],
                        'error_minus_2sigma': row['Error-2σ'],
                        'error_plus_2sigma': row['Error+2σ'],
                        'mean_check': row['Mean Check'],
                        'sigma_check': row['Mean±2σ Check']
                    })
                
                chart_info['groups'].append({
                    'group_id': channel_id,
                    'group_label': f"CH{channel_id}",
                    'color': group_colors[channel_id],
                    'items': items_data
                })
        
        charts_data.append(chart_info)
    
    # Generate HTML
    html_content = generate_comparison_html(charts_data, unit, legend_items, group_colors, files_info, group_by, sample_ids, channel_ids, report_name)
    
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✓ Comparison HTML report saved to {html_file}")
    return html_file


def generate_comparison_html(charts_data, unit, legend_items, group_colors, files_info, group_by, sample_ids, channel_ids, report_name='Comparison'):
    """Generate the HTML content for the comparison report."""
    
    report_generated_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Build legend HTML based on grouping mode
    if group_by == 'sample':
        legend_title = "Equipment Samples"
        legend_html = ''.join([
            f'<div class="legend-item"><span class="legend-color" style="background:{group_colors[s]}"></span>{s}</div>'
            for s in sample_ids
        ])
    else:
        legend_title = "Channels"
        legend_html = ''.join([
            f'<div class="legend-item"><span class="legend-color" style="background:{group_colors[ch]}"></span>CH{ch}</div>'
            for ch in channel_ids
        ])
    
    # Build charts HTML using Plotly
    charts_html = ""
    charts_js = ""
    
    for idx, chart in enumerate(charts_data):
        chart_id = f"chart_{idx}"
        
        charts_html += f'''
        <div class="chart-container">
            <div class="chart-title {chart['io_type'].lower()}">{chart['title']}</div>
            <div class="chart-wrapper" id="{chart_id}"></div>
        </div>
        '''
        
        # Build Plotly traces
        traces = []
        
        # Build x positions and labels
        x_positions = []
        x_labels = []
        pos = 0
        group_boundaries = []
        
        for group_info in chart['groups']:
            group_start = pos
            for item in group_info['items']:
                x_positions.append(pos)
                if group_by == 'sample':
                    x_labels.append(item['label'])  # equipment number (xxxxx-xxx)
                else:
                    x_labels.append(item['label'][0:-3])  # channel number (CHx)
                pos += 1
            group_boundaries.append({
                'start': group_start, 
                'end': pos - 1, 
                'group_id': group_info['group_id'],
                'color': group_info['color']
            })
            pos += 0.5  # Gap between groups
        
        x_min = -0.5
        x_max = pos - 0.5
        
        tolerance = chart['tolerance']
        
        # Upper limit line
        traces.append({
            'x': [x_min, x_max],
            'y': [tolerance, tolerance],
            'mode': 'lines',
            'name': f'+Tolerance ({tolerance:.6f})',
            'line': {'color': '#8B0000', 'width': 2, 'dash': 'dash'},
            'hoverinfo': 'name+y'
        })
        
        # Zero reference line
        traces.append({
            'x': [x_min, x_max],
            'y': [0, 0],
            'mode': 'lines',
            'name': 'Zero Error (Reference)',
            'line': {'color': '#2E7D32', 'width': 2},
            'hoverinfo': 'name+y'
        })
        
        # Lower limit line
        traces.append({
            'x': [x_min, x_max],
            'y': [-tolerance, -tolerance],
            'mode': 'lines',
            'name': f'-Tolerance ({-tolerance:.6f})',
            'line': {'color': '#8B0000', 'width': 2, 'dash': 'dash'},
            'hoverinfo': 'name+y'
        })
        
        # Add data points for each group
        pos = 0
        for group_info in chart['groups']:
            color = group_info['color']
            
            for item in group_info['items']:
                hover_text = f"{item['sample_id']} CH{item['channel']}"
                
                # Error point (diamond)
                traces.append({
                    'x': [pos],
                    'y': [item['error']],
                    'mode': 'markers',
                    'name': hover_text,
                    'marker': {'symbol': 'diamond', 'size': 10, 'color': color},
                    'hovertemplate': f"{hover_text}<br>Error: %{{y:.6f}}<br>Check: {item['mean_check']}<extra></extra>",
                    'showlegend': False
                })
                
                # Error bars (-2σ to +2σ)
                traces.append({
                    'x': [pos, pos],
                    'y': [item['error_minus_2sigma'], item['error_plus_2sigma']],
                    'mode': 'lines',
                    'line': {'color': color, 'width': 1},
                    'showlegend': False,
                    'hoverinfo': 'skip'
                })
                
                # -2σ marker
                traces.append({
                    'x': [pos],
                    'y': [item['error_minus_2sigma']],
                    'mode': 'markers',
                    'marker': {'symbol': 'line-ew', 'size': 8, 'color': color, 'line': {'color': color, 'width': 2}},
                    'hovertemplate': f"{hover_text}<br>Error-2σ: %{{y:.6f}}<extra></extra>",
                    'showlegend': False
                })
                
                # +2σ marker
                traces.append({
                    'x': [pos],
                    'y': [item['error_plus_2sigma']],
                    'mode': 'markers',
                    'marker': {'symbol': 'line-ew', 'size': 8, 'color': color, 'line': {'color': color, 'width': 2}},
                    'hovertemplate': f"{hover_text}<br>Error+2σ: %{{y:.6f}}<br>±2σ Check: {item['sigma_check']}<extra></extra>",
                    'showlegend': False
                })
                
                pos += 1
            pos += 0.5  # Gap between groups
        
        # Create shapes for group backgrounds
        shapes = []
        for boundary in group_boundaries:
            shapes.append({
                'type': 'rect',
                'xref': 'x',
                'yref': 'paper',
                'x0': boundary['start'] - 0.4,
                'x1': boundary['end'] + 0.4,
                'y0': 0,
                'y1': 1,
                'fillcolor': boundary['color'],
                'opacity': 0.1,
                'line': {'width': 0}
            })
        
        # X-axis title based on grouping mode
        if group_by == 'sample':
            x_axis_title = 'Channel'
        else:
            x_axis_title = 'Equipment Sample'
        
        # Convert to JSON for JavaScript
        import json
        traces_json = json.dumps(traces)
        shapes_json = json.dumps(shapes)
        
        if group_by == 'sample':
            charts_js += f'''
            Plotly.newPlot('{chart_id}', {traces_json}, {{
                title: null,
                xaxis: {{
                    title: '{x_axis_title}',
                    tickmode: 'array',
                    tickvals: {json.dumps(x_positions)},
                    ticktext: {json.dumps(x_labels)},
                }},
                yaxis: {{
                    title: 'Error from Reference [{unit}]',
                    zeroline: true
                }},
                shapes: {shapes_json},
                showlegend: false,
                hovermode: 'closest',
                plot_bgcolor: 'white',
                paper_bgcolor: 'white',
                margin: {{l: 60, r: 20, t: 30, b: 80}},
                autosize: true
            }}, {{responsive: true}});
            '''
        else:
            charts_js += f'''
            Plotly.newPlot('{chart_id}', {traces_json}, {{
                title: null,
                xaxis: {{
                    title: '{x_axis_title}',
                    tickmode: 'array',
                    tickvals: {json.dumps(x_positions)},
                    ticktext: {json.dumps(x_labels)},
                    tickangle: -45,
                    tickfont: {{size: 8}}
                }},
                yaxis: {{
                    title: 'Error from Reference [{unit}]',
                    zeroline: true
                }},
                shapes: {shapes_json},
                showlegend: false,
                hovermode: 'closest',
                plot_bgcolor: 'white',
                paper_bgcolor: 'white',
                margin: {{l: 60, r: 20, t: 30, b: 80}},
                autosize: true
            }}, {{responsive: true}});
            '''
    
    # Grouping mode description
    if group_by == 'sample':
        grouping_desc = "Data is grouped by <strong>equipment sample</strong>. Each color represents a different sample, and channels within each sample are grouped together."
    else:
        grouping_desc = "Data is grouped by <strong>channel number</strong>. Each color represents a different channel, and equipment samples within each channel are grouped together."
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_name} - Cross-Equipment Comparison Report</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f7fa;
            color: #333;
            line-height: 1.6;
        }}
        .header {{
            background: linear-gradient(135deg, #1F4E78 0%, #2E7D32 50%, #E8F0E8 100%);
            color: white;
            padding: 30px 40px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        .header p {{
            opacity: 0.9;
            font-size: 14px;
        }}
        .container {{
            max-width: 1600px;
            margin: 0 auto;
            padding: 20px 30px;
        }}
        .summary-section {{
            background: white;
            border-radius: 10px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        .summary-section h2 {{
            color: #1F4E78;
            margin-bottom: 15px;
            font-size: 18px;
        }}
        .sample-legend {{
            display: flex;
            flex-wrap: wrap;
            gap: 15px;
        }}
        .legend-item {{
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 14px;
        }}
        .legend-color {{
            width: 20px;
            height: 20px;
            border-radius: 4px;
        }}
        .charts-section {{
            background: white;
            border-radius: 10px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        .charts-section h2 {{
            color: #1F4E78;
            margin-bottom: 20px;
            font-size: 18px;
        }}
        .chart-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(700px, 1fr));
            gap: 25px;
        }}
        .chart-container {{
            background: #fafafa;
            border-radius: 8px;
            padding: 15px;
            border: 1px solid #e9ecef;
        }}
        .chart-title {{
            font-size: 16px;
            font-weight: 600;
            margin-bottom: 10px;
            padding: 8px 12px;
            border-radius: 0 4px 4px 0;
        }}
        .chart-title.input {{
            color: #2E7D32;
            background: linear-gradient(90deg, #e8f5e9 0%, transparent 100%);
            border-left: 4px solid #2E7D32;
        }}
        .chart-title.output {{
            color: #1F4E78;
            background: linear-gradient(90deg, #e0f0ff 0%, transparent 100%);
            border-left: 4px solid #1F4E78;
        }}
        .chart-wrapper {{
            width: 100%;
            height: 450px;
        }}
        .info-box {{
            background: #e8f4fd;
            border-left: 4px solid #1F4E78;
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 0 8px 8px 0;
        }}
        .info-box h3 {{
            color: #1F4E78;
            margin-bottom: 8px;
            font-size: 14px;
        }}
        .info-box p {{
            color: #555;
            font-size: 13px;
            margin: 0;
        }}
        .limit-legend {{
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
            padding: 10px 15px;
            background: #f8f9fa;
            border-radius: 8px;
            font-size: 13px;
            margin-bottom: 20px;
        }}
        .limit-legend-item {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .limit-line {{
            width: 30px;
            height: 3px;
        }}
        .limit-marker {{
            width: 12px;
            height: 12px;
        }}
        @media (max-width: 768px) {{
            .chart-grid {{
                grid-template-columns: 1fr;
            }}
            .header {{
                padding: 20px;
            }}
            .header h1 {{
                font-size: 22px;
            }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 {report_name} - Cross-Equipment Comparison Report</h1>
        <p>Comparing {len(sample_ids)} equipment samples | Generated: {report_generated_str}</p>
    </div>
    
    <div class="container">
        <div class="info-box">
            <h3>📌 Understanding This Report</h3>
            <p>This report shows the <strong>error from reference value</strong> (normalized data) for each measurement. 
            The Y-axis represents the difference between the measured mean and the reference value. 
            Zero means perfect accuracy. The dashed lines show the ±Tolerance limits.</p>
            <p style="margin-top: 8px;">{grouping_desc}</p>
        </div>
        
        <div class="summary-section">
            <h2> {legend_title}</h2>
            <div class="sample-legend">
                {legend_html}
            </div>
        </div>
        
        <div class="charts-section">
            <h2>📈 Comparison Charts</h2>
            <div class="limit-legend">
                <div class="limit-legend-item">
                    <div class="limit-line" style="background: #8B0000; border-style: dashed;"></div>
                    <span>±Tolerance Limits</span>
                </div>
                <div class="limit-legend-item">
                    <div class="limit-line" style="background: #2E7D32;"></div>
                    <span>Zero Error (Reference)</span>
                </div>
                <div class="limit-legend-item">
                    <div class="limit-marker" style="background: #4472C4; transform: rotate(45deg);"></div>
                    <span>Mean Error</span>
                </div>
                <div class="limit-legend-item">
                    <div class="limit-line" style="background: #4472C4; height: 2px;"></div>
                    <span>Mean ± 2σ Error</span>
                </div>
            </div>
            <div class="chart-grid">
                {charts_html}
            </div>
        </div>
    </div>
    
    <script>
        {charts_js}
        
        // Resize charts on window resize
        window.addEventListener('resize', function() {{
            const charts = document.querySelectorAll('.chart-wrapper');
            charts.forEach(function(chart) {{
                Plotly.Plots.resize(chart);
            }});
        }});
    </script>
</body>
</html>
'''
    return html


@app.route('/comparison-results')
def comparison_results():
    """Results page for cross-equipment comparison."""
    if 'comparison_output' not in session:
        return redirect(url_for('comparison_report'))
    
    return render_template('comparison_results.html',
                          output_files=session.get('comparison_output'))


@app.route('/api/upload', methods=['POST'])
def upload_files():
    """Handle file uploads."""
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files selected'}), 400
    
    session_folder = get_session_folder()
    
    # Clear previous uploads
    for old_file in session_folder.iterdir():
        old_file.unlink()
    
    uploaded_files = []
    csv_count = 0
    txt_count = 0
    
    for file in files:
        if file.filename:
            # Sanitize filename
            filename = Path(file.filename).name
            if filename.lower().endswith('.csv'):
                csv_count += 1
            elif filename.lower().endswith('.txt'):
                txt_count += 1
            else:
                continue  # Skip non-csv/txt files
            
            file_path = session_folder / filename
            file.save(str(file_path))
            uploaded_files.append(filename)
    
    if not uploaded_files:
        return jsonify({'error': 'No valid CSV or TXT files found'}), 400
    
    # Detect unit from files
    unit = get_unit_from_files(str(session_folder))
    
    # Scan for measurement types in text files
    txt_files = list(session_folder.glob('*.txt'))
    file_measurement_types = {}
    
    for txt_file in txt_files:
        types = scan_text_file_for_measurement_types(txt_file)
        if types and len(types) > 1:
            file_measurement_types[txt_file.name] = sorted(list(types))
    
    # Extract test value configurations
    test_configs = extract_test_configs(session_folder, unit)
    
    # Store in session
    session['files_info'] = {
        'count': len(uploaded_files),
        'csv_count': csv_count,
        'txt_count': txt_count,
        'unit': unit,
        'filenames': uploaded_files
    }
    session['measurement_types'] = file_measurement_types
    session['test_configs'] = test_configs
    
    return jsonify({
        'success': True,
        'files_count': len(uploaded_files),
        'csv_count': csv_count,
        'txt_count': txt_count,
        'unit': unit,
        'measurement_types': file_measurement_types,
        'test_configs': test_configs
    })


def extract_test_configs(input_dir, unit):
    """Extract unique (test_value, range_setting, io_type) tuples from filenames."""
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    
    test_configs = []
    seen = set()
    
    # CSV files are Output devices
    for csv_file in csv_files:
        value, _, channel, range_setting = parse_filename(csv_file.name)
        if value is not None and channel is not None:
            key = (value, range_setting, 'Output')
            if key not in seen:
                seen.add(key)
                test_configs.append({
                    'test_value': value,
                    'range_setting': range_setting if range_setting else 'N/A',
                    'io_type': 'Output',
                    'reference': value,
                    'tolerance': 0.015
                })
    
    # TXT files are Input devices
    for txt_file in txt_files:
        value, _, _, range_setting = parse_filename(txt_file.name)
        if value is not None:
            key = (value, range_setting, 'Input')
            if key not in seen:
                seen.add(key)
                test_configs.append({
                    'test_value': value,
                    'range_setting': range_setting if range_setting else 'N/A',
                    'io_type': 'Input',
                    'reference': value,
                    'tolerance': 0.015
                })
    
    # Sort by test value, then I/O type, then range
    test_configs.sort(key=lambda x: (x['test_value'], x['io_type'], x['range_setting']))
    
    return test_configs


@app.route('/configure')
def configure():
    """Configuration page for measurement types and tolerances."""
    if 'files_info' not in session:
        return redirect(url_for('equipment_report'))
    
    return render_template('configure.html',
                          files_info=session.get('files_info'),
                          measurement_types=session.get('measurement_types', {}),
                          test_configs=session.get('test_configs', []))


@app.route('/api/process', methods=['POST'])
def process_files():
    """Process uploaded files with user configuration."""
    if 'files_info' not in session:
        return jsonify({'error': 'No files uploaded'}), 400
    
    data = request.json
    measurement_type_selections = data.get('measurement_types', {})
    user_configs = data.get('configs', [])
    equipment_number = data.get('equipment_number', '')  # e.g., "50920-001"
    
    session_folder = get_session_folder()
    output_folder = get_output_folder()
    
    # Convert user configs to the expected format
    user_inputs = {}
    for config in user_configs:
        test_value = config['test_value']
        range_setting = config['range_setting'] if config['range_setting'] != 'N/A' else None
        io_type = config['io_type']
        
        range_input = config.get('range_input', config['range_setting'])
        final_range = None if range_input == 'N/A' or range_input == '' else range_input
        
        user_inputs[(test_value, range_setting, io_type)] = {
            'range': final_range,
            'reference': float(config['reference']),
            'tolerance': float(config['tolerance'])
        }
    
    # Convert measurement type selections keys
    full_path_selections = {}
    for filename, selected_type in measurement_type_selections.items():
        full_path = str(session_folder / filename)
        full_path_selections[full_path] = selected_type
    
    unit = session['files_info']['unit']
    
    try:
        output_file, html_file, equipment_name = process_measurement_files(
            input_dir=str(session_folder),
            output_dir=str(output_folder),
            user_inputs=user_inputs,
            unit=unit,
            measurement_type_selections=full_path_selections,
            equipment_number=equipment_number
        )
        
        # Store output file paths in session
        session['output_files'] = {
            'excel': os.path.basename(output_file) if output_file else None,
            'html': os.path.basename(html_file) if html_file else None,
            'equipment_name': equipment_name
        }
        
        return jsonify({
            'success': True,
            'excel_file': os.path.basename(output_file) if output_file else None,
            'html_file': os.path.basename(html_file) if html_file else None
        })
        
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


def process_measurement_files(input_dir, output_dir, user_inputs, unit, measurement_type_selections=None, equipment_number=None):
    """
    Process all CSV and TXT files in the input directory and compile results into Excel.
    Modified version of the original process_files function to support web output.
    Returns: (output_file, html_file, equipment_name)
    """
    dir_name = Path(input_dir).name
    if not dir_name:
        dir_name = Path(input_dir).resolve().name
    
    results = []
    
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    
    total_files = len(csv_files) + len(txt_files)
    if total_files == 0:
        raise ValueError(f"No CSV or TXT files found in {input_dir}")
    
    # Extract equipment model from the first file (e.g., "VIO2004")
    all_data_files = csv_files + txt_files
    equipment_model = None
    if all_data_files:
        first_file = all_data_files[0]
        equipment_model = extract_equipment_name(first_file.name)
    
    # Combine model and number for full equipment name
    # e.g., "VIO2004_50920-001" or just "VIO2004" if no number provided
    if equipment_number:
        equipment_name = f"{equipment_model}_{equipment_number}" if equipment_model else equipment_number
    else:
        equipment_name = equipment_model if equipment_model else dir_name
    
    # Use equipment name for output filename
    output_base_name = equipment_name
    
    # Generate versioned output filename in output directory
    base_output_file = os.path.join(output_dir, f'{output_base_name}.xlsx')
    output_file = get_versioned_filename(base_output_file)
    
    # Get timestamp of the first data file (for report header)
    if all_data_files:
        first_file = min(all_data_files, key=lambda f: f.stat().st_mtime)
        data_file_timestamp = datetime.fromtimestamp(first_file.stat().st_mtime)
    else:
        data_file_timestamp = None
    
    # Process CSV files (output data)
    for csv_file in csv_files:
        value, file_unit, channel, range_setting = parse_filename(csv_file.name)
        
        if value is None or channel is None:
            continue
        
        try:
            df = pd.read_csv(csv_file)
            
            measurement_col = None
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(keyword in col_lower for keyword in ['voltage', 'vdc', 'resistance', 'ohm', 'current', 'adc', 'measurement']):
                    measurement_col = col
                    break
            
            if measurement_col is None:
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) == 0:
                    continue
                measurement_col = numeric_cols[-1]
            
            measurements = df[measurement_col].dropna()
            
            if len(measurements) == 0:
                continue
            
            result = {
                'Channel': channel,
                'I/O Type': 'Output',
                'Range Setting': range_setting if range_setting else 'N/A',
                f'Test Value [{unit}]': value,
                f'Mean [{unit}]': measurements.mean(),
                f'StdDev [{unit}]': measurements.std(),
                f'Min [{unit}]': measurements.min(),
                f'Max [{unit}]': measurements.max(),
                'Samples': len(measurements),
                '_range_key': range_setting
            }
            
            results.append(result)
            
        except Exception as e:
            continue
    
    # Process TXT files (input data)
    for txt_file in txt_files:
        value, file_unit, channel_from_name, range_setting = parse_filename(txt_file.name)
        
        if value is None:
            continue
        
        try:
            # Get selected measurement type for this file
            selected_type = None
            if measurement_type_selections and str(txt_file) in measurement_type_selections:
                selected_type = measurement_type_selections[str(txt_file)]
            
            # Parse the text file
            channel_data = parse_text_file(txt_file, selected_measurement_type=selected_type,
                                          channel_from_filename=channel_from_name)
            
            if not channel_data:
                continue
            
            # Process each channel found in the file
            for channel, measurements in channel_data.items():
                if len(measurements) == 0:
                    continue
                
                measurements = pd.Series(measurements).dropna()
                
                if len(measurements) == 0:
                    continue
                
                result = {
                    'Channel': channel,
                    'I/O Type': 'Input',
                    'Range Setting': range_setting if range_setting else 'N/A',
                    f'Test Value [{unit}]': value,
                    f'Mean [{unit}]': measurements.mean(),
                    f'StdDev [{unit}]': measurements.std(),
                    f'Min [{unit}]': measurements.min(),
                    f'Max [{unit}]': measurements.max(),
                    'Samples': len(measurements),
                    '_range_key': range_setting
                }
                
                results.append(result)
            
        except Exception as e:
            continue
    
    if not results:
        raise ValueError("No valid results to save")
    
    # Create DataFrame and sort
    df_results = pd.DataFrame(results)
    df_results = df_results.sort_values(['Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]'])
    
    # Add reference value, tolerance, and limits columns
    if user_inputs:
        def get_user_config(row):
            test_val = row[f'Test Value [{unit}]']
            range_key = row['_range_key']
            io_type = row['I/O Type']
            key = (test_val, range_key, io_type)
            return user_inputs.get(key, {})
        
        df_results['_config'] = df_results.apply(get_user_config, axis=1)
        
        df_results[f'Reference Value [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('reference', np.nan) if x else np.nan
        )
        df_results[f'Tolerance [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('tolerance', np.nan) if x else np.nan
        )
        
        df_results['Range Setting'] = df_results.apply(
            lambda row: row['_config'].get('range', row['Range Setting'])
                        if row['_config'] and row['_config'].get('range') is not None
                        else row['Range Setting'] if row['Range Setting'] != 'N/A' else 'N/A',
            axis=1
        )
        
        # Calculate limits using reference value
        df_results[f'Lower Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] - df_results[f'Tolerance [{unit}]']
        df_results[f'Upper Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] + df_results[f'Tolerance [{unit}]']
        
        df_results['Mean Check'] = df_results.apply(
            lambda row: 'PASS' if row[f'Lower Limit [{unit}]'] <= row[f'Mean [{unit}]'] <= row[f'Upper Limit [{unit}]'] else 'FAIL',
            axis=1
        )
        
        df_results['Mean±2σ Check'] = df_results.apply(
            lambda row: 'PASS' if (
                row[f'Lower Limit [{unit}]'] <= (row[f'Mean [{unit}]'] - 2*row[f'StdDev [{unit}]']) and
                (row[f'Mean [{unit}]'] + 2*row[f'StdDev [{unit}]']) <= row[f'Upper Limit [{unit}]']
            ) else 'FAIL',
            axis=1
        )
        
        df_results = df_results.drop(columns=['_config', '_range_key'])
        
        column_order = [
            'Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]',
            f'Reference Value [{unit}]', f'Tolerance [{unit}]',
            f'Lower Limit [{unit}]', f'Upper Limit [{unit}]',
            f'Mean [{unit}]', f'StdDev [{unit}]', f'Min [{unit}]', f'Max [{unit}]',
            'Samples', 'Mean Check', 'Mean±2σ Check'
        ]
        df_results = df_results[column_order]
    else:
        df_results = df_results.drop(columns=['_range_key'])
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Test Results', index=False)
        
        worksheet = writer.sheets['Test Results']
        worksheet.auto_filter.ref = worksheet.dimensions
        
        if user_inputs:
            numeric_cols = [4, 5, 6, 7, 8, 9, 10, 11, 12]
            samples_col = 13
            pass_fail_cols = [14, 15]
        else:
            numeric_cols = [4, 5, 6, 7, 8]
            samples_col = 9
            pass_fail_cols = []
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column in numeric_cols:
                    cell.number_format = '0.000000'
                elif cell.column == samples_col:
                    cell.number_format = '0'
                elif cell.column in pass_fail_cols:
                    if cell.value == 'PASS':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        cell.font = Font(color='006100', bold=True)
                    elif cell.value == 'FAIL':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        cell.font = Font(color='9C0006', bold=True)
        
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    html_file = None
    if user_inputs:
        color_assignments = create_tolerance_charts(output_file, df_results, unit)
        
        # Apply channel colors to Test Results sheet
        if color_assignments:
            apply_channel_colors_to_results(output_file, df_results, unit, color_assignments)
        
        # Generate interactive HTML report
        html_file = create_html_report(output_file, df_results, unit, data_file_timestamp, equipment_name)
    
    return output_file, html_file, equipment_name


@app.route('/results')
def results():
    """Results page showing generated reports."""
    if 'output_files' not in session:
        return redirect(url_for('equipment_report'))
    
    return render_template('results.html',
                          output_files=session.get('output_files'))


@app.route('/download/<filename>')
def download_file(filename):
    """Download generated files."""
    output_folder = get_output_folder()
    file_path = output_folder / filename
    
    if not file_path.exists():
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(str(file_path), as_attachment=True)


@app.route('/generate-pdf/<html_filename>')
def generate_pdf(html_filename):
    """Generate PDF from the report by reading data from Excel and creating static charts."""
    output_folder = get_output_folder()
    html_path = output_folder / html_filename
    
    if not html_path.exists():
        return jsonify({'error': 'HTML file not found'}), 404
    
    try:
        import plotly.graph_objects as go
        import plotly.io as pio
        
        # Generate PDF filename
        pdf_filename = html_filename.replace('.html', '.pdf')
        pdf_path = output_folder / pdf_filename
        
        # Find the corresponding Excel file
        excel_filename = html_filename.replace('.html', '.xlsx')
        excel_path = output_folder / excel_filename
        
        if not excel_path.exists():
            # Try to find any Excel file in the output folder with similar name
            base_name = html_filename.replace('_report.html', '').replace('.html', '')
            possible_excel = list(output_folder.glob(f'{base_name}*.xlsx'))
            if possible_excel:
                excel_path = possible_excel[0]
            else:
                return jsonify({'error': 'Excel file not found. Cannot generate PDF.'}), 404
        
        # Read data from Excel
        df = pd.read_excel(excel_path, sheet_name='Test Results')
        
        # Get unit from column names
        unit = 'V'
        for col in df.columns:
            if 'Test Value [' in col:
                unit = col.split('[')[1].split(']')[0]
                break
        
        # Handle NaN in Range Setting
        df['Range Setting'] = df['Range Setting'].fillna('N/A')
        
        # Get unique test value + range + I/O type combinations
        unique_combos = df.groupby([f'Test Value [{unit}]', 'Range Setting', 'I/O Type']).size().reset_index()
        
        from utils import CHANNEL_COLORS_HEX
        
        # Create figures for each chart
        figs = []
        
        for _, combo in unique_combos.iterrows():
            test_value = combo[f'Test Value [{unit}]']
            range_setting = combo['Range Setting']
            io_type = combo['I/O Type']
            
            mask = (
                (df[f'Test Value [{unit}]'] == test_value) &
                (df['Range Setting'] == range_setting) &
                (df['I/O Type'] == io_type)
            )
            chart_data = df[mask].sort_values('Channel')
            
            if len(chart_data) == 0:
                continue
            
            fig = go.Figure()
            
            # Get limits
            reference = chart_data[f'Reference Value [{unit}]'].iloc[0]
            tolerance = chart_data[f'Tolerance [{unit}]'].iloc[0]
            upper_limit = chart_data[f'Upper Limit [{unit}]'].iloc[0]
            lower_limit = chart_data[f'Lower Limit [{unit}]'].iloc[0]
            
            channels = chart_data['Channel'].tolist()
            x_range = [-0.5, len(channels) - 0.5]
            
            # Upper limit
            fig.add_trace(go.Scatter(
                x=x_range, y=[upper_limit, upper_limit],
                mode='lines', name=f'Upper Limit ({upper_limit:.4f})',
                line=dict(color='#8B0000', width=2, dash='dash')
            ))
            
            # Reference line
            fig.add_trace(go.Scatter(
                x=x_range, y=[reference, reference],
                mode='lines', name=f'Reference ({reference:.4f})',
                line=dict(color='#2E7D32', width=2)
            ))
            
            # Lower limit
            fig.add_trace(go.Scatter(
                x=x_range, y=[lower_limit, lower_limit],
                mode='lines', name=f'Lower Limit ({lower_limit:.4f})',
                line=dict(color='#8B0000', width=2, dash='dash')
            ))
            
            # Add channel data
            x_labels = []
            for i, (_, row) in enumerate(chart_data.iterrows()):
                color = CHANNEL_COLORS_HEX[i % len(CHANNEL_COLORS_HEX)]
                ch = int(row['Channel'])
                x_labels.append(f'CH{ch}')
                
                mean_val = row[f'Mean [{unit}]']
                std_val = row[f'StdDev [{unit}]']
                mean_minus_2sigma = mean_val - 2 * std_val
                mean_plus_2sigma = mean_val + 2 * std_val
                
                # Mean marker (diamond)
                fig.add_trace(go.Scatter(
                    x=[i], y=[mean_val],
                    mode='markers',
                    name=f'CH{ch}',
                    marker=dict(symbol='diamond', size=12, color=color),
                    hovertemplate=f'CH{ch}<br>Mean: {mean_val:.6f}<extra></extra>'
                ))
                
                # Error bars (-2σ to +2σ)
                fig.add_trace(go.Scatter(
                    x=[i, i], y=[mean_minus_2sigma, mean_plus_2sigma],
                    mode='lines',
                    line=dict(color=color, width=2),
                    showlegend=False,
                    hoverinfo='skip'
                ))
                
                # Endpoints
                fig.add_trace(go.Scatter(
                    x=[i], y=[mean_minus_2sigma],
                    mode='markers',
                    marker=dict(symbol='line-ew', size=8, color=color, line=dict(width=2, color=color)),
                    showlegend=False,
                    hovertemplate=f'CH{ch}<br>Mean-2σ: {mean_minus_2sigma:.6f}<extra></extra>'
                ))
                fig.add_trace(go.Scatter(
                    x=[i], y=[mean_plus_2sigma],
                    mode='markers',
                    marker=dict(symbol='line-ew', size=8, color=color, line=dict(width=2, color=color)),
                    showlegend=False,
                    hovertemplate=f'CH{ch}<br>Mean+2σ: {mean_plus_2sigma:.6f}<extra></extra>'
                ))
            
            # Chart title
            range_display = f", Range: {range_setting}" if range_setting != 'N/A' else ""
            title = f"Test: {test_value} {unit}{range_display} ({io_type})"
            
            fig.update_layout(
                title=dict(text=title, font=dict(size=14)),
                xaxis=dict(
                    title='Channel',
                    tickmode='array',
                    tickvals=list(range(len(x_labels))),
                    ticktext=x_labels
                ),
                yaxis=dict(title=f'Value [{unit}]'),
                showlegend=True,
                legend=dict(font=dict(size=9)),
                width=900,
                height=500,
                template='plotly_white',
                margin=dict(l=60, r=150, t=50, b=50)
            )
            
            figs.append(fig)
        
        if not figs:
            return jsonify({'error': 'No charts to generate'}), 400
        
        # Export to PDF
        if len(figs) == 1:
            figs[0].write_image(str(pdf_path), format='pdf')
        else:
            # For multiple charts, combine into single PDF
            import tempfile
            try:
                from pypdf import PdfWriter, PdfReader
            except ImportError:
                from PyPDF2 import PdfWriter, PdfReader
            
            writer = PdfWriter()
            temp_files = []
            
            for i, fig in enumerate(figs):
                temp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
                temp_files.append(temp_pdf.name)
                fig.write_image(temp_pdf.name, format='pdf')
                temp_pdf.close()
            
            # Merge all PDFs
            for temp_file in temp_files:
                reader = PdfReader(temp_file)
                for page in reader.pages:
                    writer.add_page(page)
            
            with open(pdf_path, 'wb') as output:
                writer.write(output)
            
            # Cleanup temp files
            for temp_file in temp_files:
                os.unlink(temp_file)
        
        return send_file(str(pdf_path), as_attachment=True, download_name=pdf_filename)
        
    except ImportError as e:
        return jsonify({'error': f'PDF generation requires additional packages (kaleido, pypdf). Install with: pip install kaleido pypdf. Error: {str(e)}'}), 500
    except Exception as e:
        import traceback
        return jsonify({'error': f'PDF generation failed: {str(e)}', 'traceback': traceback.format_exc()}), 500


@app.route('/view/<filename>')
def view_file(filename):
    """View HTML report in browser."""
    output_folder = get_output_folder()
    file_path = output_folder / filename
    
    if not file_path.exists():
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(str(file_path))


@app.route('/api/save-config', methods=['POST'])
def save_config():
    """Save configuration to JSON file."""
    data = request.json
    output_folder = get_output_folder()
    
    config_file = output_folder / 'test_config.json'
    
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
    
    return jsonify({
        'success': True,
        'filename': 'test_config.json'
    })


@app.route('/api/load-config', methods=['POST'])
def load_config():
    """Load configuration from uploaded JSON file."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        content = file.read().decode('utf-8')
        config_data = json.loads(content)
        return jsonify({
            'success': True,
            'config': config_data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/reset')
def reset_session():
    """Reset session and clean up files."""
    if 'session_id' in session:
        session_folder = UPLOAD_FOLDER / session['session_id']
        output_folder = OUTPUT_FOLDER / session['session_id']
        
        if session_folder.exists():
            shutil.rmtree(session_folder)
        if output_folder.exists():
            shutil.rmtree(output_folder)
    
    session.clear()
    return jsonify({'success': True})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
