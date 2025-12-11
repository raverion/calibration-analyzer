from pathlib import Path
from datetime import datetime

from utils import PLOTLY_AVAILABLE, CHANNEL_COLORS_HEX

if PLOTLY_AVAILABLE:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

'''
def create_tolerance_charts(excel_file, df_results, unit):
    """
    Create tolerance charts showing limits, reference value, mean, and mean±2σ for each 
    test value + range setting combination.
    
    Returns: Dictionary mapping (channel, io_type, test_value, range_setting) to color
    """
    print("\nCreating Tolerance charts...")
    
    wb = load_workbook(excel_file)
    
    if 'Tolerance Charts' in wb.sheetnames:
        del wb['Tolerance Charts']
    chart_sheet = wb.create_sheet('Tolerance Charts')
    
    # Get unique combinations
    unique_combinations = df_results.groupby([f'Test Value [{unit}]', 'Range Setting', 'I/O Type']).size().reset_index()
    unique_combinations = unique_combinations.sort_values([f'Test Value [{unit}]', 'Range Setting', 'I/O Type'])
    
    charts_per_row = 2
    chart_width = 26
    chart_height = 36
    
    # Pleasant, muted color palette for channels (same color for mean, +2σ, -2σ)
    channel_colors = [
        '4472C4',  # Muted blue
        'C45B5B',  # Muted red
        '70AD47',  # Muted green
        'ED7D31',  # Muted orange
        '7B7B7B',  # Gray
        '9E5ECE',  # Muted purple
        '43A6A2',  # Muted teal
        'C4A24E',  # Muted gold
        '5B9BC4',  # Steel blue
        'A85B5B',  # Dusty rose
        '5BAF7B',  # Sea green
        'C47B4E',  # Terracotta
        '6B6BAF',  # Muted indigo
        '8B6BAF',  # Muted violet
        '4EAFAF',  # Turquoise
        'AF8B4E',  # Bronze
    ]
    
    # Track color assignments for each channel within each chart context
    # Key: (channel, io_type, test_value, range_setting) -> color
    color_assignments = {}
    
    chart_idx = 0
    
    from openpyxl.styles import Font, Alignment, PatternFill
    
    for _, combo_row in unique_combinations.iterrows():
        test_value = combo_row[f'Test Value [{unit}]']
        range_setting = combo_row['Range Setting']
        io_type = combo_row['I/O Type']
        
        mask = (
            (df_results[f'Test Value [{unit}]'] == test_value) &
            (df_results['Range Setting'] == range_setting) &
            (df_results['I/O Type'] == io_type)
        )
        test_data = df_results[mask].copy()
        test_data = test_data.sort_values('Channel')
        
        if len(test_data) == 0:
            continue
        
        col_offset = (chart_idx % charts_per_row) * chart_width
        row_offset = (chart_idx // charts_per_row) * chart_height
        
        range_display = f", Range: {range_setting}" if range_setting != 'N/A' else ""
        chart_title = f"Test: {test_value} {unit}{range_display} ({io_type})"
        
        title_row = row_offset + 1
        chart_sheet.cell(title_row, col_offset + 1).value = chart_title
        title_cell = chart_sheet.cell(title_row, col_offset + 1)
        title_cell.font = Font(bold=True, size=11, color='1F4E78')
        title_cell.alignment = Alignment(horizontal='left')
        
        channels = test_data['Channel'].tolist()
        num_channels = len(channels)
        lower_limits = test_data[f'Lower Limit [{unit}]'].tolist()
        upper_limits = test_data[f'Upper Limit [{unit}]'].tolist()
        reference_values = test_data[f'Reference Value [{unit}]'].tolist()
        means = test_data[f'Mean [{unit}]'].tolist()
        lower_2sigma = (test_data[f'Mean [{unit}]'] - 2*test_data[f'StdDev [{unit}]']).tolist()
        upper_2sigma = (test_data[f'Mean [{unit}]'] + 2*test_data[f'StdDev [{unit}]']).tolist()
        mean_checks = test_data['Mean Check'].tolist()
        mean_2sigma_checks = test_data['Mean±2σ Check'].tolist()
        
        # Get limit values (same for all channels in this chart)
        ref_val = reference_values[0]
        ll_val = lower_limits[0]
        ul_val = upper_limits[0]
        
        data_start_row = row_offset + 3
        data_start_col = col_offset + 1
        
        # Write headers
        headers = ['Channel', 'Lower Limit', 'Reference', 'Upper Limit', 'Mean', 'Mean-2σ', 'Mean+2σ', 'Mean Check', 'Mean±2σ Check']
        for h_idx, header in enumerate(headers):
            chart_sheet.cell(data_start_row, data_start_col + h_idx).value = header
            chart_sheet.cell(data_start_row, data_start_col + h_idx).font = Font(bold=True, size=9)
        
        # Write data with color-coded fonts
        for i, channel in enumerate(channels):
            row = data_start_row + i + 1
            color = channel_colors[i % len(channel_colors)]
            
            # Store color assignment
            color_assignments[(channel, io_type, test_value, range_setting)] = color
            
            # Apply color to all cells in this row
            cell_font = Font(size=9, color=color)
            
            chart_sheet.cell(row, data_start_col, channel).font = cell_font
            chart_sheet.cell(row, data_start_col + 1, lower_limits[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 1).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 2, reference_values[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 2).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 3, upper_limits[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 3).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 4, means[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 4).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 5, lower_2sigma[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 5).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 6, upper_2sigma[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 6).number_format = '0.000000'
            
            # Mean Check column with PASS/FAIL formatting
            mean_check_cell = chart_sheet.cell(row, data_start_col + 7, mean_checks[i])
            if mean_checks[i] == 'PASS':
                mean_check_cell.font = Font(size=9, color='006100', bold=True)
                mean_check_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                mean_check_cell.font = Font(size=9, color='9C0006', bold=True)
                mean_check_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Mean±2σ Check column with PASS/FAIL formatting
            mean_2sigma_check_cell = chart_sheet.cell(row, data_start_col + 8, mean_2sigma_checks[i])
            if mean_2sigma_checks[i] == 'PASS':
                mean_2sigma_check_cell.font = Font(size=9, color='006100', bold=True)
                mean_2sigma_check_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                mean_2sigma_check_cell.font = Font(size=9, color='9C0006', bold=True)
                mean_2sigma_check_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Create scatter chart
        chart = ScatterChart()
        chart.title = chart_title
        chart.style = 10
        
        # Set axis titles (non-bold is default for axis titles)
        chart.x_axis.title = "Channel"
        chart.y_axis.title = f"Measurement [{unit}]"
        
        # Chart size - make it larger for better visibility
        chart.height = 14
        chart.width = 18
        
        # Remove legend
        chart.legend = None
        
        # Remove gridlines
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
        
        # Calculate Y-axis limits with padding
        all_values = lower_limits + upper_limits + reference_values + means + lower_2sigma + upper_2sigma
        y_min = min(all_values)
        y_max = max(all_values)
        y_range = y_max - y_min if y_max != y_min else abs(y_max) * 0.1 or 0.1
        y_padding = y_range * 0.20  # 20% padding
        
        chart.y_axis.scaling.min = y_min - y_padding
        chart.y_axis.scaling.max = y_max + y_padding
        
        # Set X-axis to properly scale for the number of channels
        x_min_val = min(channels)
        x_max_val = max(channels)
        x_padding_left = max(0.8, (x_max_val - x_min_val) * 0.1)
        x_padding_right = max(0.8, (x_max_val - x_min_val) * 0.1)
        chart.x_axis.scaling.min = x_min_val - x_padding_left
        chart.x_axis.scaling.max = x_max_val + x_padding_right
        chart.x_axis.majorUnit = 1  # Show each channel number
        
        # Reference for x values (channel numbers)
        xvalues = Reference(chart_sheet, min_col=data_start_col, min_row=data_start_row+1, max_row=data_start_row+num_channels)
        
        # Lower limit line (dark red, dashed)
        lower_series = Series(Reference(chart_sheet, min_col=data_start_col+1, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                            xvalues, title="Lower Limit")
        lower_series.marker = Marker('none')
        lower_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="8B0000", w=12700, prstDash='dash'))
        chart.series.append(lower_series)
        
        # Reference value line (dark green, solid)
        ref_series = Series(Reference(chart_sheet, min_col=data_start_col+2, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                           xvalues, title="Reference")
        ref_series.marker = Marker('none')
        ref_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="2E7D32", w=12700))
        chart.series.append(ref_series)
        
        # Upper limit line (dark red, dashed)
        upper_series = Series(Reference(chart_sheet, min_col=data_start_col+3, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                            xvalues, title="Upper Limit")
        upper_series.marker = Marker('none')
        upper_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="8B0000", w=12700, prstDash='dash'))
        chart.series.append(upper_series)
        
        # Add series for each channel with consistent colors
        for i, channel in enumerate(channels):
            color = channel_colors[i % len(channel_colors)]
            channel_row = data_start_row + 1 + i
            
            # Mean (diamond marker - smaller size)
            mean_series = Series(Reference(chart_sheet, min_col=data_start_col+4, min_row=channel_row, max_row=channel_row), 
                               Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                               title=f"CH{channel} Mean")
            mean_series.marker = Marker('diamond', size=6)
            mean_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color))
            mean_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(mean_series)
            
            # Mean-2σ (horizontal line marker - thinner)
            lower_2s_series = Series(Reference(chart_sheet, min_col=data_start_col+5, min_row=channel_row, max_row=channel_row), 
                                   Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                                   title=f"CH{channel} -2σ")
            lower_2s_series.marker = Marker('dash', size=8)
            lower_2s_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color, w=12700))
            lower_2s_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(lower_2s_series)
            
            # Mean+2σ (horizontal line marker - thinner)
            upper_2s_series = Series(Reference(chart_sheet, min_col=data_start_col+6, min_row=channel_row, max_row=channel_row), 
                                   Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                                   title=f"CH{channel} +2σ")
            upper_2s_series.marker = Marker('dash', size=8)
            upper_2s_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color, w=12700))
            upper_2s_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(upper_2s_series)
        
        # Position chart (moved further right to accommodate check columns)
        chart_cell = chart_sheet.cell(row_offset + 4, col_offset + 12)
        chart.anchor = chart_cell.coordinate
        chart_sheet.add_chart(chart)
        
        print(f"  Created chart for {chart_title}")
        chart_idx += 1
    
    wb.save(excel_file)
    print("✓ Tolerance charts added to workbook")
    
    return color_assignments
'''
def create_html_report(output_file, df_results, unit, data_file_timestamp=None, equipment_name=None):
    """
    Create an interactive HTML report using Plotly with tolerance charts and data tables.
    
    Parameters:
    - output_file: Path to the output Excel file
    - df_results: DataFrame with results
    - unit: Measurement unit
    - data_file_timestamp: Optional datetime of the first data file (for "Data collected on")
    - equipment_name: Optional equipment model name for report title
    """
    if not PLOTLY_AVAILABLE:
        print("Warning: Plotly not available. Skipping HTML report generation.")
        print("  Install with: pip install plotly")
        return None
    
    print("\nCreating interactive HTML report...")
    
    # Use equipment name if provided, otherwise use file stem
    report_name = equipment_name if equipment_name else Path(output_file).stem
    
    # Generate HTML filename
    html_file = output_file.replace('.xlsx', '_report.html')
    
    # Pleasant, muted color palette for channels (matching Excel charts)
    channel_colors = [
        '#4472C4',  # Muted blue
        '#C45B5B',  # Muted red
        '#70AD47',  # Muted green
        '#ED7D31',  # Muted orange
        '#7B7B7B',  # Gray
        '#9E5ECE',  # Muted purple
        '#43A6A2',  # Muted teal
        '#C4A24E',  # Muted gold
        '#5B9BC4',  # Steel blue
        '#A85B5B',  # Dusty rose
        '#5BAF7B',  # Sea green
        '#C47B4E',  # Terracotta
        '#6B6BAF',  # Muted indigo
        '#8B6BAF',  # Muted violet
        '#4EAFAF',  # Turquoise
        '#AF8B4E',  # Bronze
    ]
    
    # Get unique combinations for charts
    unique_combinations = df_results.groupby([f'Test Value [{unit}]', 'Range Setting', 'I/O Type']).size().reset_index()
    unique_combinations = unique_combinations.sort_values([f'Test Value [{unit}]', 'Range Setting', 'I/O Type'])
    
    num_charts = len(unique_combinations)
    
    # Create figures list
    figures_html = []
    
    for _, combo_row in unique_combinations.iterrows():
        test_value = combo_row[f'Test Value [{unit}]']
        range_setting = combo_row['Range Setting']
        io_type = combo_row['I/O Type']
        
        mask = (
            (df_results[f'Test Value [{unit}]'] == test_value) &
            (df_results['Range Setting'] == range_setting) &
            (df_results['I/O Type'] == io_type)
        )
        test_data = df_results[mask].copy()
        test_data = test_data.sort_values('Channel')
        
        if len(test_data) == 0:
            continue
        
        channels = test_data['Channel'].tolist()
        lower_limits = test_data[f'Lower Limit [{unit}]'].tolist()
        upper_limits = test_data[f'Upper Limit [{unit}]'].tolist()
        reference_values = test_data[f'Reference Value [{unit}]'].tolist()
        means = test_data[f'Mean [{unit}]'].tolist()
        stddevs = test_data[f'StdDev [{unit}]'].tolist()
        lower_2sigma = [m - 2*s for m, s in zip(means, stddevs)]
        upper_2sigma = [m + 2*s for m, s in zip(means, stddevs)]
        mean_checks = test_data['Mean Check'].tolist()
        mean_2sigma_checks = test_data['Mean±2σ Check'].tolist()
        
        # Get limit values (same for all channels)
        ref_val = reference_values[0]
        ll_val = lower_limits[0]
        ul_val = upper_limits[0]
        
        range_display = f", Range: {range_setting}" if range_setting != 'N/A' else ""
        chart_title = f"Test: {test_value} {unit}{range_display} ({io_type})"
        
        # Create figure
        fig = go.Figure()
        
        # Add limit lines (horizontal lines across all channels)
        x_range = [min(channels) - 0.5, max(channels) + 0.5]
        
        # Lower limit line (dashed red)
        fig.add_trace(go.Scatter(
            x=x_range,
            y=[ll_val, ll_val],
            mode='lines',
            name=f'Lower Limit ({ll_val:.6f})',
            line=dict(color='#8B0000', width=2, dash='dash'),
            hoverinfo='name+y'
        ))
        
        # Reference line (solid green)
        fig.add_trace(go.Scatter(
            x=x_range,
            y=[ref_val, ref_val],
            mode='lines',
            name=f'Reference ({ref_val:.6f})',
            line=dict(color='#2E7D32', width=2),
            hoverinfo='name+y'
        ))
        
        # Upper limit line (dashed red)
        fig.add_trace(go.Scatter(
            x=x_range,
            y=[ul_val, ul_val],
            mode='lines',
            name=f'Upper Limit ({ul_val:.6f})',
            line=dict(color='#8B0000', width=2, dash='dash'),
            hoverinfo='name+y'
        ))
        
        # Add data points for each channel
        for i, channel in enumerate(channels):
            color = channel_colors[i % len(channel_colors)]
            
            # Mean point (diamond)
            fig.add_trace(go.Scatter(
                x=[channel],
                y=[means[i]],
                mode='markers',
                name=f'CH{channel} Mean',
                marker=dict(
                    symbol='diamond',
                    size=12,
                    color=color,
                    line=dict(color=color, width=1)
                ),
                hovertemplate=f'CH{channel}<br>Mean: %{{y:.6f}}<br>Check: {mean_checks[i]}<extra></extra>'
            ))
            
            # Mean-2σ point (line marker)
            fig.add_trace(go.Scatter(
                x=[channel],
                y=[lower_2sigma[i]],
                mode='markers',
                name=f'CH{channel} -2σ',
                marker=dict(
                    symbol='line-ew',
                    size=10,
                    color=color,
                    line=dict(color=color, width=3)
                ),
                hovertemplate=f'CH{channel}<br>Mean-2σ: %{{y:.6f}}<extra></extra>',
                showlegend=False
            ))
            
            # Mean+2σ point (line marker)
            fig.add_trace(go.Scatter(
                x=[channel],
                y=[upper_2sigma[i]],
                mode='markers',
                name=f'CH{channel} +2σ',
                marker=dict(
                    symbol='line-ew',
                    size=10,
                    color=color,
                    line=dict(color=color, width=3)
                ),
                hovertemplate=f'CH{channel}<br>Mean+2σ: %{{y:.6f}}<br>±2σ Check: {mean_2sigma_checks[i]}<extra></extra>',
                showlegend=False
            ))
            
            # Add vertical line connecting -2σ to +2σ
            fig.add_trace(go.Scatter(
                x=[channel, channel],
                y=[lower_2sigma[i], upper_2sigma[i]],
                mode='lines',
                line=dict(color=color, width=1),
                showlegend=False,
                hoverinfo='skip'
            ))
        
        # Calculate Y-axis range
        all_values = lower_limits + upper_limits + reference_values + means + lower_2sigma + upper_2sigma
        y_min = min(all_values)
        y_max = max(all_values)
        y_range_val = y_max - y_min if y_max != y_min else abs(y_max) * 0.1 or 0.1
        y_padding = y_range_val * 0.20
        
        # Update layout - title moved outside chart, legends hidden by default
        fig.update_layout(
            title=None,  # Title will be added as external HTML element
            xaxis=dict(
                title='Channel',
                tickmode='linear',
                tick0=min(channels),
                dtick=1,
                autorange=True  # Enable autoscale
            ),
            yaxis=dict(
                title=f'Measurement [{unit}]',
                autorange=True  # Enable autoscale
            ),
            legend=dict(
                orientation='h',
                yanchor='bottom',
                y=1.0,
                xanchor='center',
                x=0.5,
                font=dict(size=10),
                bgcolor='rgba(255,255,255,0.9)',
                bordercolor='#e9ecef',
                borderwidth=1
            ),
            showlegend=False,  # Legends hidden by default
            hovermode='closest',
            plot_bgcolor='white',
            paper_bgcolor='white',
            margin=dict(l=60, r=20, t=30, b=50),
            autosize=True
        )
        
        # Add gridlines
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='#E0E0E0')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='#E0E0E0')
        
        # Store chart with its title and io_type
        figures_html.append({
            'title': chart_title,
            'io_type': io_type,
            'html': fig.to_html(full_html=False, include_plotlyjs=False, config={'responsive': True})
        })
    
    # Create Deviation Summary Charts (all channels, all test values in one chart)
    # Group by I/O type AND Range Setting to create separate deviation charts
    deviation_charts = []
    
    # Get all unique channels for consistent color assignment across all charts
    all_channels = sorted(df_results['Channel'].unique())
    channel_color_map = {ch: channel_colors[i % len(channel_colors)] for i, ch in enumerate(all_channels)}
    
    # Get unique combinations of I/O type and Range Setting
    io_range_combinations = df_results.groupby(['I/O Type', 'Range Setting']).size().reset_index()
    io_range_combinations = io_range_combinations.sort_values(['I/O Type', 'Range Setting'])
    
    for _, combo in io_range_combinations.iterrows():
        io_type = combo['I/O Type']
        range_setting = combo['Range Setting']
        
        # Filter data for this combination
        mask = (df_results['I/O Type'] == io_type) & (df_results['Range Setting'] == range_setting)
        combo_data = df_results[mask].copy()
        
        if len(combo_data) == 0:
            continue
        
        # Get unique channels and test values for this combination
        channels = sorted(combo_data['Channel'].unique())
        test_values = sorted(combo_data[f'Test Value [{unit}]'].unique())
        
        # Create figure
        fig = go.Figure()
        
        # Add a line for each channel
        for channel in channels:
            ch_data = combo_data[combo_data['Channel'] == channel].copy()
            ch_data = ch_data.sort_values(f'Test Value [{unit}]')
            
            x_vals = ch_data[f'Test Value [{unit}]'].tolist()
            # Calculate deviation (Mean - Reference)
            deviations = (ch_data[f'Mean [{unit}]'] - ch_data[f'Reference Value [{unit}]']).tolist()
            means = ch_data[f'Mean [{unit}]'].tolist()
            refs = ch_data[f'Reference Value [{unit}]'].tolist()
            
            color = channel_color_map[channel]
            
            # Add line with markers
            fig.add_trace(go.Scatter(
                x=x_vals,
                y=deviations,
                mode='lines+markers',
                name=f'CH{channel}',
                line=dict(color=color, width=2),
                marker=dict(color=color, size=8, symbol='diamond'),
                hovertemplate=(
                    f'<b>Channel {channel}</b><br>'
                    f'Test Value: %{{x}} {unit}<br>'
                    f'Deviation: %{{y:.6f}} {unit}<br>'
                    f'Mean: %{{customdata[0]:.6f}} {unit}<br>'
                    f'Reference: %{{customdata[1]:.6f}} {unit}<extra></extra>'
                ),
                customdata=list(zip(means, refs))
            ))
        
        # Add zero reference line
        fig.add_hline(
            y=0,
            line=dict(color='#2E7D32', width=2),
            annotation_text="Zero Deviation",
            annotation_position="bottom right"
        )
        
        # Update layout
        io_label = "Input" if io_type == "Input" else "Output"
        range_label = f" (Range: {range_setting})" if range_setting and range_setting != 'N/A' else ""
        chart_title = f'Deviation Summary - {io_label}{range_label}'
        
        fig.update_layout(
            title=None,
            xaxis_title=f'Test Value [{unit}]',
            yaxis_title=f'Deviation [{unit}]',
            font=dict(family='Segoe UI', size=11),
            legend=dict(
                orientation='h',
                yanchor='bottom',
                y=1.02,
                xanchor='center',
                x=0.5,
                bgcolor='rgba(255,255,255,0.9)',
                bordercolor='#e9ecef',
                borderwidth=1
            ),
            showlegend=True,
            hovermode='closest',
            plot_bgcolor='white',
            paper_bgcolor='white',
            margin=dict(l=60, r=20, t=50, b=50),
            autosize=True
        )
        
        # Add gridlines
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='#E0E0E0')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='#E0E0E0', zeroline=True, zerolinecolor='#2E7D32', zerolinewidth=2)
        
        deviation_charts.append({
            'title': chart_title,
            'io_type': io_type,
            'html': fig.to_html(full_html=False, include_plotlyjs=False, config={'responsive': True})
        })
    
    # Create summary statistics table
    summary_pass = (df_results['Mean Check'] == 'PASS').sum()
    summary_fail = (df_results['Mean Check'] == 'FAIL').sum()
    summary_2s_pass = (df_results['Mean±2σ Check'] == 'PASS').sum()
    summary_2s_fail = (df_results['Mean±2σ Check'] == 'FAIL').sum()
    
    # Get unique values for filters
    unique_channels = sorted(df_results['Channel'].unique())
    unique_ranges = sorted(df_results['Range Setting'].unique())
    unique_test_values = sorted(df_results[f'Test Value [{unit}]'].unique())
    unique_io_types = sorted(df_results['I/O Type'].unique())
    
    # Generate timestamp information
    report_generated_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if data_file_timestamp:
        data_collected_str = data_file_timestamp.strftime("%Y-%m-%d %H:%M:%S")
    else:
        data_collected_str = "Unknown"
    
    # Build HTML document
    html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Measurement Report - {report_name}</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f7fa;
            color: #333;
            line-height: 1.6;
        }}
        .header {{
            background: linear-gradient(135deg, #5C2D91 0%, #9B59B6 50%, #E8E0F0 100%);
            color: white;
            padding: 30px 40px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        .header p {{
            opacity: 0.9;
            font-size: 14px;
        }}
        .container {{
            width: 100%;
            padding: 20px 30px;
        }}
        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 20px;
            margin: 20px 0;
        }}
        @media (max-width: 1200px) {{
            .summary-cards {{
                grid-template-columns: repeat(3, 1fr);
            }}
        }}
        @media (max-width: 768px) {{
            .summary-cards {{
                grid-template-columns: repeat(2, 1fr);
            }}
        }}
        .card {{
            background: white;
            border-radius: 8px;
            padding: 12px 15px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            text-align: center;
        }}
        .card h3 {{
            font-size: 11px;
            color: #666;
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .card .value {{
            font-size: 24px;
            font-weight: bold;
        }}
        .card.pass .value {{
            color: #2E7D32;
        }}
        .card.fail .value {{
            color: #C00000;
        }}
        .card.neutral .value {{
            color: #1F4E78;
        }}
        .section {{
            background: white;
            border-radius: 10px;
            margin: 20px 0;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            overflow: hidden;
        }}
        .section-header {{
            background: #f8f9fa;
            padding: 15px 20px;
            border-bottom: 1px solid #e9ecef;
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .section-header:hover {{
            background: #e9ecef;
        }}
        .section-header h2 {{
            font-size: 18px;
            color: #1F4E78;
        }}
        .section-header .toggle {{
            font-size: 20px;
            color: #666;
        }}
        .section-content {{
            padding: 20px;
        }}
        .chart-grid {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 25px;
        }}
        @media (max-width: 1400px) {{
            .chart-grid {{
                grid-template-columns: 1fr;
            }}
        }}
        .chart-container {{
            background: #fafafa;
            border-radius: 8px;
            padding: 15px;
            border: 1px solid #e9ecef;
            min-height: 400px;
        }}
        .chart-title {{
            font-size: 16px;
            font-weight: 600;
            margin-bottom: 10px;
            padding: 8px 12px;
            border-radius: 0 4px 4px 0;
        }}
        .chart-title.input {{
            color: #2E7D32;
            background: linear-gradient(90deg, #e8f5e9 0%, transparent 100%);
            border-left: 4px solid #2E7D32;
        }}
        .chart-title.output {{
            color: #1F4E78;
            background: linear-gradient(90deg, #e0f0ff 0%, transparent 100%);
            border-left: 4px solid #1F4E78;
        }}
        .chart-wrapper {{
            width: 100%;
            height: 380px;
        }}
        .chart-wrapper > div {{
            width: 100% !important;
            height: 100% !important;
        }}
        .deviation-chart-grid {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 25px;
        }}
        @media (max-width: 1400px) {{
            .deviation-chart-grid {{
                grid-template-columns: 1fr;
            }}
        }}
        .deviation-chart-container {{
            background: #fafafa;
            border-radius: 8px;
            padding: 15px;
            border: 1px solid #e9ecef;
            min-height: 450px;
        }}
        .deviation-chart {{
            height: 400px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        th, td {{
            padding: 10px 12px;
            text-align: left;
            border-bottom: 1px solid #e9ecef;
        }}
        th {{
            background: #f8f9fa;
            font-weight: 600;
            color: #1F4E78;
            position: sticky;
            top: 0;
            white-space: nowrap;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        .pass {{
            background-color: #C6EFCE;
            color: #006100;
            font-weight: bold;
            text-align: center;
            border-radius: 4px;
        }}
        .fail {{
            background-color: #FFC7CE;
            color: #9C0006;
            font-weight: bold;
            text-align: center;
            border-radius: 4px;
        }}
        .table-wrapper {{
            max-height: 500px;
            overflow-y: auto;
            border: 1px solid #e9ecef;
            border-radius: 8px;
        }}
        .chart-controls {{
            display: flex;
            align-items: center;
            gap: 20px;
            margin-bottom: 15px;
            flex-wrap: wrap;
        }}
        .toggle-legend-btn {{
            padding: 8px 16px;
            background: #1F4E78;
            color: white;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            transition: background 0.2s;
            display: flex;
            align-items: center;
            gap: 6px;
        }}
        .toggle-legend-btn:hover {{
            background: #2E7D32;
        }}
        .toggle-legend-btn.legends-hidden {{
            background: #666;
        }}
        .filter-bar {{
            display: flex;
            gap: 10px;
            margin-bottom: 15px;
            flex-wrap: wrap;
            align-items: center;
        }}
        .filter-group {{
            display: flex;
            align-items: center;
            gap: 6px;
            background: #f8f9fa;
            padding: 6px 10px;
            border-radius: 6px;
        }}
        .filter-bar label {{
            font-weight: 500;
            color: #666;
            font-size: 13px;
            white-space: nowrap;
        }}
        .filter-bar select, .filter-bar input {{
            padding: 6px 10px;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 13px;
            min-width: 100px;
        }}
        .filter-bar input {{
            min-width: 150px;
        }}
        .filter-bar select:focus, .filter-bar input:focus {{
            outline: none;
            border-color: #1F4E78;
        }}
        .clear-filters-btn {{
            padding: 6px 12px;
            background: #e9ecef;
            color: #666;
            border: 1px solid #ddd;
            border-radius: 4px;
            cursor: pointer;
            font-size: 13px;
            transition: all 0.2s;
        }}
        .clear-filters-btn:hover {{
            background: #ddd;
            color: #333;
        }}
        .filter-count {{
            font-size: 12px;
            color: #666;
            padding: 4px 8px;
            background: #e9ecef;
            border-radius: 4px;
        }}
        .legend-info {{
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
            padding: 10px 15px;
            background: #f8f9fa;
            border-radius: 8px;
            font-size: 13px;
        }}
        .legend-item {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .legend-line {{
            width: 30px;
            height: 3px;
        }}
        .legend-marker {{
            width: 12px;
            height: 12px;
        }}
        .collapsible {{
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.5s ease-out;
        }}
        .collapsible.active {{
            max-height: none;
            overflow: visible;
        }}
        @media (max-width: 768px) {{
            .chart-grid {{
                grid-template-columns: 1fr;
            }}
            .header {{
                padding: 20px;
            }}
            .header h1 {{
                font-size: 22px;
            }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 {report_name} - Measurement Analysis Report</h1>
        <p>Report generated on: {report_generated_str} | Data collected on: {data_collected_str}</p>
    </div>
    
    <div class="container">
        <!-- Summary Cards -->
        <div class="summary-cards">
            <div class="card neutral">
                <h3>Total Tests</h3>
                <div class="value">{len(df_results)}</div>
            </div>
            <div class="card pass">
                <h3>Mean Check Pass</h3>
                <div class="value">{summary_pass}</div>
            </div>
            <div class="card fail">
                <h3>Mean Check Fail</h3>
                <div class="value">{summary_fail}</div>
            </div>
            <div class="card pass">
                <h3>±2σ Check Pass</h3>
                <div class="value">{summary_2s_pass}</div>
            </div>
            <div class="card fail">
                <h3>±2σ Check Fail</h3>
                <div class="value">{summary_2s_fail}</div>
            </div>
        </div>
        
        <!-- Charts Section -->
        <div class="section">
            <div class="section-header" onclick="toggleSection('charts-section')">
                <h2>📈 Tolerance Charts</h2>
                <span class="toggle" id="charts-section-toggle">▼</span>
            </div>
            <div class="section-content collapsible active" id="charts-section">
                <div class="chart-controls">
                    <button class="toggle-legend-btn legends-hidden" onclick="toggleAllLegends()">
                        <span id="legend-btn-icon">👁️‍🗨️</span> Toggle Legends
                    </button>
                    <div class="legend-info">
                        <div class="legend-item">
                            <div class="legend-line" style="background: #8B0000; border-style: dashed;"></div>
                            <span>Upper/Lower Limits</span>
                        </div>
                        <div class="legend-item">
                            <div class="legend-line" style="background: #2E7D32;"></div>
                            <span>Reference Value</span>
                        </div>
                        <div class="legend-item">
                            <div class="legend-marker" style="background: #4472C4; transform: rotate(45deg);"></div>
                            <span>Mean</span>
                        </div>
                        <div class="legend-item">
                            <div class="legend-line" style="background: #4472C4; height: 2px;"></div>
                            <span>Mean ± 2σ</span>
                        </div>
                    </div>
                </div>
                <div class="chart-grid">
'''
    
    # Add each chart with external title (color based on I/O type)
    for i, chart_data in enumerate(figures_html):
        io_class = 'input' if chart_data['io_type'] == 'Input' else 'output'
        html_content += f'''
                    <div class="chart-container">
                        <div class="chart-title {io_class}">{chart_data['title']}</div>
                        <div class="chart-wrapper">
                            {chart_data['html']}
                        </div>
                    </div>
'''
    
    html_content += '''
                </div>
            </div>
        </div>
        
        <!-- Deviation Summary Section -->
        <div class="section">
            <div class="section-header" onclick="toggleSection('deviation-section')">
                <h2>📉 Deviation Summary</h2>
                <span class="toggle" id="deviation-section-toggle">▼</span>
            </div>
            <div class="section-content collapsible active" id="deviation-section">
                <p style="color: #666; margin-bottom: 15px; font-size: 0.9rem;">
                    Shows the deviation (Mean - Reference) for all channels across all test values. 
                    The green line indicates zero deviation.
                </p>
                <div class="deviation-chart-grid">
'''
    
    # Add deviation charts
    for chart_data in deviation_charts:
        io_class = 'input' if chart_data['io_type'] == 'Input' else 'output'
        html_content += f'''
                    <div class="deviation-chart-container">
                        <div class="chart-title {io_class}">{chart_data['title']}</div>
                        <div class="chart-wrapper deviation-chart">
                            {chart_data['html']}
                        </div>
                    </div>
'''
    
    html_content += '''
                </div>
            </div>
        </div>
        
        <!-- Data Table Section -->
        <div class="section">
            <div class="section-header" onclick="toggleSection('data-section')">>
                <h2>📋 Detailed Results</h2>
                <span class="toggle" id="data-section-toggle">▼</span>
            </div>
            <div class="section-content collapsible active" id="data-section">
                <div class="filter-bar">
                    <div class="filter-group">
                        <label>Channel:</label>
                        <select id="channel-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add channel options
    for ch in unique_channels:
        html_content += f'                            <option value="{ch}">{ch}</option>\n'
    
    html_content += '''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>I/O Type:</label>
                        <select id="io-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add I/O type options
    for io in unique_io_types:
        html_content += f'                            <option value="{io}">{io}</option>\n'
    
    html_content += '''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Range:</label>
                        <select id="range-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add range options
    for rng in unique_ranges:
        html_content += f'                            <option value="{rng}">{rng}</option>\n'
    
    html_content += f'''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Test Value:</label>
                        <select id="testvalue-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add test value options
    for tv in unique_test_values:
        html_content += f'                            <option value="{tv}">{tv} {unit}</option>\n'
    
    html_content += '''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Status:</label>
                        <select id="status-filter" onchange="filterTable()">
                            <option value="all">All</option>
                            <option value="pass">Pass Only</option>
                            <option value="fail">Fail Only</option>
                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Search:</label>
                        <input type="text" id="search-input" placeholder="Search..." onkeyup="filterTable()">
                    </div>
                    <button class="clear-filters-btn" onclick="clearFilters()">Clear All</button>
                    <span class="filter-count" id="filter-count"></span>
                </div>
                <div class="table-wrapper">
                    <table id="results-table">
                        <thead>
                            <tr>
                                <th>Channel</th>
                                <th>I/O Type</th>
                                <th>Range</th>
'''
    
    html_content += f'''
                                <th>Test Value [{unit}]</th>
                                <th>Reference [{unit}]</th>
                                <th>Tolerance [{unit}]</th>
                                <th>Lower Limit [{unit}]</th>
                                <th>Upper Limit [{unit}]</th>
                                <th>Mean [{unit}]</th>
                                <th>StdDev [{unit}]</th>
                                <th>Min [{unit}]</th>
                                <th>Max [{unit}]</th>
                                <th>Samples</th>
                                <th>Mean Check</th>
                                <th>Mean±2σ Check</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    # Add table rows
    for _, row in df_results.iterrows():
        mean_class = 'pass' if row['Mean Check'] == 'PASS' else 'fail'
        sigma_class = 'pass' if row['Mean±2σ Check'] == 'PASS' else 'fail'
        
        html_content += f'''
                            <tr>
                                <td>{row['Channel']}</td>
                                <td>{row['I/O Type']}</td>
                                <td>{row['Range Setting']}</td>
                                <td>{row[f'Test Value [{unit}]']:.6f}</td>
                                <td>{row[f'Reference Value [{unit}]']:.6f}</td>
                                <td>{row[f'Tolerance [{unit}]']:.6f}</td>
                                <td>{row[f'Lower Limit [{unit}]']:.6f}</td>
                                <td>{row[f'Upper Limit [{unit}]']:.6f}</td>
                                <td>{row[f'Mean [{unit}]']:.6f}</td>
                                <td>{row[f'StdDev [{unit}]']:.6f}</td>
                                <td>{row[f'Min [{unit}]']:.6f}</td>
                                <td>{row[f'Max [{unit}]']:.6f}</td>
                                <td>{row['Samples']}</td>
                                <td class="{mean_class}">{row['Mean Check']}</td>
                                <td class="{sigma_class}">{row['Mean±2σ Check']}</td>
                            </tr>
'''
    
    html_content += '''
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>
    
    <script>
        let legendsVisible = false;
        
        function toggleSection(sectionId) {
            const section = document.getElementById(sectionId);
            const toggle = document.getElementById(sectionId + '-toggle');
            section.classList.toggle('active');
            toggle.textContent = section.classList.contains('active') ? '▼' : '▶';
        }
        
        function toggleAllLegends() {
            legendsVisible = !legendsVisible;
            const charts = document.querySelectorAll('.chart-wrapper .plotly-graph-div');
            const btn = document.querySelector('.toggle-legend-btn');
            const icon = document.getElementById('legend-btn-icon');
            
            charts.forEach(function(chart) {
                Plotly.relayout(chart, { showlegend: legendsVisible });
            });
            
            if (legendsVisible) {
                btn.classList.remove('legends-hidden');
                icon.textContent = '👁️';
            } else {
                btn.classList.add('legends-hidden');
                icon.textContent = '👁️‍🗨️';
            }
        }
        
        function filterTable() {
            const channelFilter = document.getElementById('channel-filter').value;
            const ioFilter = document.getElementById('io-filter').value;
            const rangeFilter = document.getElementById('range-filter').value;
            const testValueFilter = document.getElementById('testvalue-filter').value;
            const statusFilter = document.getElementById('status-filter').value;
            const searchInput = document.getElementById('search-input').value.toLowerCase();
            const table = document.getElementById('results-table');
            const rows = table.getElementsByTagName('tr');
            
            let visibleCount = 0;
            let totalCount = rows.length - 1; // Exclude header
            
            for (let i = 1; i < rows.length; i++) {
                const row = rows[i];
                const cells = row.getElementsByTagName('td');
                
                let showRow = true;
                
                // Channel filter (column 0)
                if (channelFilter !== 'all' && showRow) {
                    const channel = cells[0].textContent;
                    if (channel !== channelFilter) showRow = false;
                }
                
                // I/O Type filter (column 1)
                if (ioFilter !== 'all' && showRow) {
                    const ioType = cells[1].textContent;
                    if (ioType !== ioFilter) showRow = false;
                }
                
                // Range filter (column 2)
                if (rangeFilter !== 'all' && showRow) {
                    const range = cells[2].textContent;
                    if (range !== rangeFilter) showRow = false;
                }
                
                // Test Value filter (column 3)
                if (testValueFilter !== 'all' && showRow) {
                    const testValue = parseFloat(cells[3].textContent);
                    const filterValue = parseFloat(testValueFilter);
                    if (Math.abs(testValue - filterValue) > 0.000001) showRow = false;
                }
                
                // Status filter (columns 13 and 14)
                if (statusFilter !== 'all' && showRow) {
                    const meanCheck = cells[13].textContent;
                    const sigmaCheck = cells[14].textContent;
                    if (statusFilter === 'pass') {
                        if (meanCheck !== 'PASS' || sigmaCheck !== 'PASS') showRow = false;
                    } else if (statusFilter === 'fail') {
                        if (meanCheck !== 'FAIL' && sigmaCheck !== 'FAIL') showRow = false;
                    }
                }
                
                // Search filter
                if (searchInput && showRow) {
                    let found = false;
                    for (let j = 0; j < cells.length; j++) {
                        if (cells[j].textContent.toLowerCase().includes(searchInput)) {
                            found = true;
                            break;
                        }
                    }
                    if (!found) showRow = false;
                }
                
                row.style.display = showRow ? '' : 'none';
                if (showRow) visibleCount++;
            }
            
            // Update filter count
            const countEl = document.getElementById('filter-count');
            if (visibleCount === totalCount) {
                countEl.textContent = `Showing all ${totalCount} rows`;
            } else {
                countEl.textContent = `Showing ${visibleCount} of ${totalCount} rows`;
            }
        }
        
        function clearFilters() {
            document.getElementById('channel-filter').value = 'all';
            document.getElementById('io-filter').value = 'all';
            document.getElementById('range-filter').value = 'all';
            document.getElementById('testvalue-filter').value = 'all';
            document.getElementById('status-filter').value = 'all';
            document.getElementById('search-input').value = '';
            filterTable();
        }
        
        // Resize all Plotly charts when window resizes
        window.addEventListener('resize', function() {
            const charts = document.querySelectorAll('.chart-wrapper .plotly-graph-div');
            charts.forEach(function(chart) {
                Plotly.Plots.resize(chart);
            });
        });
        
        // Initial setup after page load
        window.addEventListener('load', function() {
            setTimeout(function() {
                const charts = document.querySelectorAll('.chart-wrapper .plotly-graph-div');
                charts.forEach(function(chart) {
                    Plotly.Plots.resize(chart);
                });
                // Initialize filter count
                filterTable();
            }, 100);
        });
    </script>
</body>
</html>
'''
    
    # Write HTML file
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✓ Interactive HTML report saved to {html_file}")
    return html_file