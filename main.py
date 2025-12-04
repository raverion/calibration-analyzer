import os
import webbrowser
import tkinter as tk
from tkinter import filedialog
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font

# Import from local modules
from parsers import (
    parse_filename,
    get_unit_from_files,
    scan_text_file_for_measurement_types,
    parse_text_file
)
from gui import select_measurement_type, get_user_inputs
from excel_charts import create_tolerance_charts, apply_channel_colors_to_results
from html_report import create_html_report
from utils import get_versioned_filename


def process_files(input_dir='.', user_inputs=None, unit='V', measurement_type_selections=None):
    """
    Process all CSV and TXT files in the input directory and compile results into Excel.
    """
    dir_name = Path(input_dir).name
    if not dir_name:
        dir_name = Path(input_dir).resolve().name
    
    # Generate versioned output filename
    base_output_file = os.path.join(input_dir, f'{dir_name}.xlsx')
    output_file = get_versioned_filename(base_output_file)
    
    results = []
    
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    
    total_files = len(csv_files) + len(txt_files)
    if total_files == 0:
        print(f"No CSV or TXT files found in {input_dir}")
        return
    
    print(f"Found {len(csv_files)} CSV files and {len(txt_files)} TXT files")
    
    # Process CSV files (output data)
    for csv_file in csv_files:
        value, file_unit, channel, range_setting = parse_filename(csv_file.name)
        
        if value is None or channel is None:
            print(f"Skipping {csv_file.name} - could not parse filename (value or channel missing)")
            continue
        
        try:
            df = pd.read_csv(csv_file)
            
            measurement_col = None
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(keyword in col_lower for keyword in ['voltage', 'vdc', 'resistance', 'ohm', 'current', 'adc', 'measurement']):
                    measurement_col = col
                    break
            
            if measurement_col is None:
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) == 0:
                    print(f"Warning: No numeric columns found in {csv_file.name}")
                    continue
                measurement_col = numeric_cols[-1]
            
            measurements = df[measurement_col].dropna()
            
            if len(measurements) == 0:
                print(f"Warning: No valid measurements in {csv_file.name}")
                continue
            
            result = {
                'Channel': channel,
                'I/O Type': 'Output',
                'Range Setting': range_setting if range_setting else 'N/A',
                f'Test Value [{unit}]': value,
                f'Mean [{unit}]': measurements.mean(),
                f'StdDev [{unit}]': measurements.std(),
                f'Min [{unit}]': measurements.min(),
                f'Max [{unit}]': measurements.max(),
                'Samples': len(measurements),
                '_range_key': range_setting
            }
            
            results.append(result)
            print(f"Processed: {csv_file.name} - CH{channel}, {value}{unit}, Range:{range_setting or 'N/A'}, {len(measurements)} samples (Output)")
            
        except Exception as e:
            print(f"Error processing {csv_file.name}: {str(e)}")
            continue
    
    # Process TXT files (input data)
    for txt_file in txt_files:
        value, file_unit, channel_from_name, range_setting = parse_filename(txt_file.name)
        
        if value is None:
            print(f"Skipping {txt_file.name} - could not parse test value from filename")
            continue
        
        try:
            # Get selected measurement type for this file
            selected_type = None
            if measurement_type_selections and str(txt_file) in measurement_type_selections:
                selected_type = measurement_type_selections[str(txt_file)]
            
            # Parse the text file, passing channel from filename if available
            channel_data = parse_text_file(txt_file, selected_measurement_type=selected_type, 
                                          channel_from_filename=channel_from_name)
            
            if not channel_data:
                print(f"Warning: No valid measurements parsed from {txt_file.name}")
                continue
            
            # Process each channel found in the file
            for channel, measurements in channel_data.items():
                if len(measurements) == 0:
                    continue
                
                measurements = pd.Series(measurements).dropna()
                
                if len(measurements) == 0:
                    continue
                
                result = {
                    'Channel': channel,
                    'I/O Type': 'Input',
                    'Range Setting': range_setting if range_setting else 'N/A',
                    f'Test Value [{unit}]': value,
                    f'Mean [{unit}]': measurements.mean(),
                    f'StdDev [{unit}]': measurements.std(),
                    f'Min [{unit}]': measurements.min(),
                    f'Max [{unit}]': measurements.max(),
                    'Samples': len(measurements),
                    '_range_key': range_setting
                }
                
                results.append(result)
                type_info = f" ({selected_type})" if selected_type else ""
                print(f"Processed: {txt_file.name} - CH{channel}, {value}{unit}, Range:{range_setting or 'N/A'}, {len(measurements)} samples (Input){type_info}")
            
        except Exception as e:
            print(f"Error processing {txt_file.name}: {str(e)}")
            continue
    
    if not results:
        print("No valid results to save")
        return
    
    # Create DataFrame and sort
    df_results = pd.DataFrame(results)
    df_results = df_results.sort_values(['Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]'])
    
    # Add reference value, tolerance, and limits columns
    if user_inputs:
        def get_user_config(row):
            test_val = row[f'Test Value [{unit}]']
            range_key = row['_range_key']
            io_type = row['I/O Type']
            key = (test_val, range_key, io_type)
            return user_inputs.get(key, {})
        
        df_results['_config'] = df_results.apply(get_user_config, axis=1)
        
        df_results[f'Reference Value [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('reference', np.nan) if x else np.nan
        )
        df_results[f'Tolerance [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('tolerance', np.nan) if x else np.nan
        )
        
        df_results['Range Setting'] = df_results.apply(
            lambda row: row['_config'].get('range', row['Range Setting']) 
                        if row['_config'] and row['_config'].get('range') is not None 
                        else row['Range Setting'] if row['Range Setting'] != 'N/A' else 'N/A',
            axis=1
        )
        
        # Calculate limits using reference value
        df_results[f'Lower Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] - df_results[f'Tolerance [{unit}]']
        df_results[f'Upper Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] + df_results[f'Tolerance [{unit}]']
        
        df_results['Mean Check'] = df_results.apply(
            lambda row: 'PASS' if row[f'Lower Limit [{unit}]'] <= row[f'Mean [{unit}]'] <= row[f'Upper Limit [{unit}]'] else 'FAIL',
            axis=1
        )
        
        df_results['Mean±2σ Check'] = df_results.apply(
            lambda row: 'PASS' if (
                row[f'Lower Limit [{unit}]'] <= (row[f'Mean [{unit}]'] - 2*row[f'StdDev [{unit}]']) and
                (row[f'Mean [{unit}]'] + 2*row[f'StdDev [{unit}]']) <= row[f'Upper Limit [{unit}]']
            ) else 'FAIL',
            axis=1
        )
        
        df_results = df_results.drop(columns=['_config', '_range_key'])
        
        column_order = [
            'Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]', 
            f'Reference Value [{unit}]', f'Tolerance [{unit}]', 
            f'Lower Limit [{unit}]', f'Upper Limit [{unit}]',
            f'Mean [{unit}]', f'StdDev [{unit}]', f'Min [{unit}]', f'Max [{unit}]', 
            'Samples', 'Mean Check', 'Mean±2σ Check'
        ]
        df_results = df_results[column_order]
    else:
        df_results = df_results.drop(columns=['_range_key'])
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Test Results', index=False)
        
        worksheet = writer.sheets['Test Results']
        worksheet.auto_filter.ref = worksheet.dimensions
        
        from openpyxl.styles import PatternFill, Font
        
        if user_inputs:
            numeric_cols = [4, 5, 6, 7, 8, 9, 10, 11, 12]
            samples_col = 13
            pass_fail_cols = [14, 15]
        else:
            numeric_cols = [4, 5, 6, 7, 8]
            samples_col = 9
            pass_fail_cols = []
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column in numeric_cols:
                    cell.number_format = '0.000000'
                elif cell.column == samples_col:
                    cell.number_format = '0'
                elif cell.column in pass_fail_cols:
                    if cell.value == 'PASS':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        cell.font = Font(color='006100', bold=True)
                    elif cell.value == 'FAIL':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        cell.font = Font(color='9C0006', bold=True)
        
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    if user_inputs:
        color_assignments = create_tolerance_charts(output_file, df_results, unit)
        
        # Apply channel colors to Test Results sheet
        if color_assignments:
            apply_channel_colors_to_results(output_file, df_results, unit, color_assignments)
        
        # Generate interactive HTML report
        html_file = create_html_report(output_file, df_results, unit)
    else:
        html_file = None
    
    print(f"\n✓ Results saved to {output_file}")
    print(f"  Total entries: {len(df_results)}")
    print(f"  Channels: {sorted(df_results['Channel'].unique())}")
    print(f"  Test values: {sorted(df_results[f'Test Value [{unit}]'].unique())}")
    print(f"  I/O Types: {sorted(df_results['I/O Type'].unique())}")
    
    return output_file, html_file


if __name__ == "__main__":
    print("=" * 70)
    print("Measurement Data Compiler with Tolerance Charts")
    print("Supports CSV (Output) and TXT (Input) files with Range Settings")
    print("=" * 70)
    print("\nPlease select the directory containing data files...")
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    input_dir = filedialog.askdirectory(
        title="Select Directory with CSV/TXT Files",
        initialdir=os.getcwd()
    )
    
    root.destroy()
    
    if not input_dir:
        print("\nNo directory selected. Exiting...")
        exit(0)
    
    print(f"\nSelected directory: {input_dir}")
    
    # Determine unit from files
    unit = get_unit_from_files(input_dir)
    print(f"Detected unit: {unit}")
    
    # Scan files
    print("\nScanning data files...")
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    all_files = csv_files + txt_files
    
    if not all_files:
        print(f"Error: No CSV or TXT files found in {input_dir}")
        exit(1)
    
    # Scan text files for measurement types
    print("Analyzing text file structures...")
    file_measurement_types = {}
    for txt_file in txt_files:
        types = scan_text_file_for_measurement_types(txt_file)
        if types:
            file_measurement_types[str(txt_file)] = types
            if len(types) > 1:
                print(f"  {txt_file.name}: Found multiple measurement types: {', '.join(sorted(types))}")
            else:
                print(f"  {txt_file.name}: Found measurement type: {list(types)[0]}")
    
    # If any file has multiple measurement types, ask user to select
    measurement_type_selections = None
    if file_measurement_types:
        measurement_type_selections = select_measurement_type(file_measurement_types)
        if measurement_type_selections is None:
            print("\nMeasurement type selection cancelled. Exiting...")
            exit(0)
    
    # Extract unique (test_value, range_setting, io_type) tuples from filenames
    test_value_range_io_tuples = set()
    
    # CSV files are Output devices
    for csv_file in csv_files:
        value, _, channel, range_setting = parse_filename(csv_file.name)
        if value is not None and channel is not None:
            test_value_range_io_tuples.add((value, range_setting, 'Output'))
    
    # TXT files are Input devices
    for txt_file in txt_files:
        value, _, _, range_setting = parse_filename(txt_file.name)
        if value is not None:
            test_value_range_io_tuples.add((value, range_setting, 'Input'))
    
    if not test_value_range_io_tuples:
        print("Error: Could not parse any valid test values from filenames")
        exit(1)
    
    print(f"\nFound {len(all_files)} data files ({len(csv_files)} CSV, {len(txt_files)} TXT)")
    print(f"Found {len(test_value_range_io_tuples)} unique test value/range/IO-type combinations")
    
    # Get user inputs
    user_inputs = get_user_inputs(test_value_range_io_tuples, unit, input_dir=input_dir)
    
    if user_inputs is None:
        print("\nConfiguration input cancelled. Exiting...")
        exit(0)
    
    print("\nProcessing files...")
    
    output_file, html_file = process_files(
        input_dir=input_dir, 
        user_inputs=user_inputs, 
        unit=unit,
        measurement_type_selections=measurement_type_selections
    )
    
    print("\n" + "=" * 70)
    print(f"Output file: {output_file}")
    if html_file:
        print(f"HTML report: {html_file}")
        # Automatically open the HTML report in the default browser
        import webbrowser
        webbrowser.open('file://' + os.path.abspath(html_file))
        print("\n✓ HTML report opened in browser")